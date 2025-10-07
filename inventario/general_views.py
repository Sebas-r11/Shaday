# inventario/general_views.py
# Vistas generales, mixins y funcionalidades de apoyo para el módulo de inventario

from django.shortcuts import render, redirect
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib.auth.decorators import login_required
from django.views.generic import TemplateView
from django.urls import reverse_lazy
from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Sum, Count, Q, F
from datetime import datetime
from decimal import Decimal

# Importaciones para Excel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Importaciones locales
from .models import Producto, Categoria, Subcategoria, Stock, Bodega

# ========================================
# MIXINS DE PERMISOS Y AUTENTICACIÓN
# ========================================

class InventarioRequiredMixin(UserPassesTestMixin):
    """Mixin para verificar permisos de inventario"""
    def test_func(self):
        return hasattr(self.request.user, 'can_adjust_inventory') and self.request.user.can_adjust_inventory()


class InventarioViewMixin(UserPassesTestMixin):
    """Mixin para permitir ver inventario (incluye bodega)"""
    def test_func(self):
        return (
            (hasattr(self.request.user, 'can_adjust_inventory') and self.request.user.can_adjust_inventory()) or 
            (hasattr(self.request.user, 'can_view_inventory') and self.request.user.can_view_inventory()) or
            self.request.user.is_superuser or
            self.request.user.role in ['superadmin', 'administrador', 'bodeguero']
        )


class AdminOnlyMixin(UserPassesTestMixin):
    """Mixin para funciones que solo puede usar el administrador"""
    def test_func(self):
        return self.request.user.is_superuser or self.request.user.groups.filter(name='admin').exists()


class AdminInventarioMixin(UserPassesTestMixin):
    """Mixin para funciones estratégicas de inventario (solo administradores)"""
    def test_func(self):
        return (
            self.request.user.is_superuser or 
            self.request.user.role in ['superadmin', 'administrador']
        )


class BodegaMixin(UserPassesTestMixin):
    """Mixin para funciones de bodega (bodegueros y administradores)"""
    def test_func(self):
        return (
            self.request.user.is_superuser or 
            self.request.user.role in ['superadmin', 'administrador', 'bodeguero'] or
            (hasattr(self.request.user, 'can_adjust_inventory') and self.request.user.can_adjust_inventory())
        )


# ========================================
# VISTAS GENERALES DEL SISTEMA
# ========================================

class InventarioHomeView(InventarioViewMixin, TemplateView):
    """Vista principal del módulo de inventario"""
    template_name = 'inventario/home.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Estadísticas rápidas
        productos_activos = Producto.objects.filter(activo=True)
        
        context['estadisticas'] = {
            'total_productos': productos_activos.count(),
            'total_categorias': Categoria.objects.filter(activa=True).count(),
            'total_bodegas': Bodega.objects.filter(activa=True).count(),
            'productos_registrados': productos_activos.count(),
            'productos_activos': productos_activos.filter(activo=True).count(),
        }
        
        # Productos más recientes
        context['productos_recientes'] = productos_activos.order_by('-fecha_creacion')[:5]
        
        # Alertas simplificadas (comentado por problemas con stock_actual)
        # context['alertas'] = {
        #     'sin_stock': productos_activos.filter(stock_actual=0).count(),
        #     'stock_critico': productos_activos.filter(
        #         stock_actual__lte=F('stock_minimo')
        #     ).count(),
        # }
        
        # Últimos movimientos (si tiene permisos)
        if self.request.user.is_superuser or self.request.user.role in ['superadmin', 'administrador']:
            from .models import MovimientoInventario
            context['ultimos_movimientos'] = MovimientoInventario.objects.select_related(
                'producto', 'bodega_origen', 'bodega_destino', 'usuario'
            ).order_by('-fecha_movimiento')[:10]
        
        return context


class InventarioMenuView(InventarioViewMixin, TemplateView):
    """Vista del menú principal de inventario"""
    template_name = 'inventario/menu.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Configurar menú según permisos del usuario
        menu_items = []
        
        # Items básicos de visualización
        if self.request.user.is_authenticated:
            menu_items.extend([
                {
                    'nombre': 'Productos',
                    'url': 'inventario:producto_list',
                    'icono': 'fas fa-box',
                    'descripcion': 'Ver y gestionar productos'
                },
                {
                    'nombre': 'Categorías',
                    'url': 'inventario:categoria_list',
                    'icono': 'fas fa-tags',
                    'descripcion': 'Gestionar categorías y subcategorías'
                },
                {
                    'nombre': 'Stock',
                    'url': 'inventario:stock_list',
                    'icono': 'fas fa-warehouse',
                    'descripcion': 'Consultar stock por bodega'
                }
            ])
        
        # Items de administración
        if self.request.user.is_superuser or self.request.user.role in ['superadmin', 'administrador']:
            menu_items.extend([
                {
                    'nombre': 'Bodegas',
                    'url': 'inventario:bodega_list',
                    'icono': 'fas fa-building',
                    'descripcion': 'Gestionar bodegas'
                },
                {
                    'nombre': 'Movimientos',
                    'url': 'inventario:movimiento_list',
                    'icono': 'fas fa-exchange-alt',
                    'descripcion': 'Historial de movimientos'
                },
                {
                    'nombre': 'Reportes',
                    'url': 'inventario:reporte_inventario',
                    'icono': 'fas fa-chart-bar',
                    'descripcion': 'Reportes y estadísticas'
                }
            ])
        
        # Items de bodeguero
        if (hasattr(self.request.user, 'can_adjust_inventory') and self.request.user.can_adjust_inventory()) or \
           self.request.user.role in ['superadmin', 'administrador', 'bodeguero']:
            menu_items.extend([
                {
                    'nombre': 'Ajustar Inventario',
                    'url': 'inventario:ajuste_inventario',
                    'icono': 'fas fa-edit',
                    'descripcion': 'Realizar ajustes de inventario'
                },
                {
                    'nombre': 'Transferencias',
                    'url': 'inventario:transferencia_create',
                    'icono': 'fas fa-truck',
                    'descripcion': 'Transferir entre bodegas'
                }
            ])
        
        context['menu_items'] = menu_items
        return context


# ========================================
# FUNCIONES DE EXPORTACIÓN E IMPORTACIÓN
# ========================================

@login_required
def exportar_productos_excel(request):
    """Vista para exportar productos a Excel"""
    if not request.user.is_authenticated:
        return HttpResponse("No autorizado", status=401)
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"
    
    # Configurar estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Definir headers basándose en permisos del usuario
    if hasattr(request.user, 'can_see_costs') and request.user.can_see_costs():
        headers = [
            'Código', 'Nombre', 'Categoría', 'Subcategoría', 
            'Costo Promedio', 'Precio Minorista', 'Precio Mayorista',
            'Stock Total', 'Stock Mínimo', 'Estado', 'Fecha Creación'
        ]
    else:
        headers = [
            'Código', 'Nombre', 'Categoría', 'Subcategoría', 
            'Precio Minorista', 'Precio Mayorista',
            'Stock Total', 'Stock Mínimo', 'Estado', 'Fecha Creación'
        ]
    
    # Escribir headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Obtener productos
    productos = Producto.objects.filter(activo=True).select_related(
        'categoria', 'subcategoria'
    ).order_by('codigo')
    
    # Escribir datos
    row_num = 2
    for producto in productos:
        if hasattr(request.user, 'can_see_costs') and request.user.can_see_costs():
            row_data = [
                producto.codigo,
                producto.nombre,
                producto.categoria.nombre if producto.categoria else '',
                producto.subcategoria.nombre if producto.subcategoria else '',
                float(producto.costo_promedio) if producto.costo_promedio else 0,
                float(producto.precio_minorista) if producto.precio_minorista else 0,
                float(producto.precio_mayorista) if producto.precio_mayorista else 0,
                0,  # stock_actual no existe, usar 0
                producto.stock_minimo,
                'Activo' if producto.activo else 'Inactivo',
                producto.fecha_creacion.strftime('%d/%m/%Y %H:%M') if producto.fecha_creacion else ''
            ]
        else:
            row_data = [
                producto.codigo,
                producto.nombre,
                producto.categoria.nombre if producto.categoria else '',
                producto.subcategoria.nombre if producto.subcategoria else '',
                float(producto.precio_minorista) if producto.precio_minorista else 0,
                float(producto.precio_mayorista) if producto.precio_mayorista else 0,
                0,  # stock_actual no existe, usar 0
                producto.stock_minimo,
                'Activo' if producto.activo else 'Inactivo',
                producto.fecha_creacion.strftime('%d/%m/%Y %H:%M') if producto.fecha_creacion else ''
            ]
        
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)
        
        row_num += 1
    
    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'productos_reyes_{timestamp}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    # Guardar workbook en la respuesta
    wb.save(response)
    
    return response


@login_required
def exportar_stock_excel(request):
    """Vista para exportar reporte de stock a Excel"""
    if not (request.user.is_superuser or request.user.role in ['superadmin', 'administrador']):
        return HttpResponse("No autorizado", status=403)
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Stock por Bodega"
    
    # Configurar estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = [
        'Código Producto', 'Nombre Producto', 'Categoría', 
        'Bodega', 'Stock Actual', 'Stock Mínimo', 'Stock Máximo',
        'Estado Stock', 'Última Actualización'
    ]
    
    # Escribir headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Obtener datos de stock
    stocks = Stock.objects.select_related(
        'producto', 'producto__categoria', 'bodega'
    ).filter(
        producto__activo=True,
        bodega__activa=True
    ).order_by('producto__codigo', 'bodega__nombre')
    
    # Escribir datos
    row_num = 2
    for stock in stocks:
        # Determinar estado del stock
        if stock.cantidad == 0:
            estado = 'Sin Stock'
        elif stock.cantidad <= stock.stock_minimo:
            estado = 'Stock Crítico'
        elif stock.cantidad <= stock.stock_minimo * 1.5:
            estado = 'Stock Bajo'
        else:
            estado = 'Stock Normal'
        
        row_data = [
            stock.producto.codigo,
            stock.producto.nombre,
            stock.producto.categoria.nombre if stock.producto.categoria else '',
            stock.bodega.nombre,
            stock.cantidad,
            stock.stock_minimo,
            stock.stock_maximo,
            estado,
            stock.fecha_actualizacion.strftime('%d/%m/%Y %H:%M') if stock.fecha_actualizacion else ''
        ]
        
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)
        
        row_num += 1
    
    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'stock_bodegas_{timestamp}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    # Guardar workbook en la respuesta
    wb.save(response)
    
    return response


# ========================================
# FUNCIONES DE UTILIDAD Y AYUDA
# ========================================

@login_required
def inventario_ayuda(request):
    """Vista de ayuda del módulo de inventario"""
    context = {
        'titulo': 'Ayuda - Módulo de Inventario',
        'secciones': [
            {
                'titulo': 'Gestión de Productos',
                'contenido': 'Aprende a crear, editar y organizar productos en el sistema.',
                'enlaces': [
                    {'nombre': 'Crear Producto', 'url': 'inventario:producto_create'},
                    {'nombre': 'Lista de Productos', 'url': 'inventario:producto_list'},
                ]
            },
            {
                'titulo': 'Control de Stock',
                'contenido': 'Monitorea y ajusta el inventario en todas las bodegas.',
                'enlaces': [
                    {'nombre': 'Ver Stock', 'url': 'inventario:stock_list'},
                    {'nombre': 'Ajustar Inventario', 'url': 'inventario:ajuste_inventario'},
                ]
            },
            {
                'titulo': 'Reportes',
                'contenido': 'Genera reportes detallados del estado del inventario.',
                'enlaces': [
                    {'nombre': 'Reporte de Inventario', 'url': 'inventario:reporte_inventario'},
                    {'nombre': 'Exportar Excel', 'url': 'inventario:exportar_productos'},
                ]
            }
        ]
    }
    
    return render(request, 'inventario/ayuda.html', context)


def validar_permisos_inventario(user):
    """Función auxiliar para validar permisos de inventario"""
    if not user.is_authenticated:
        return False
    
    return (
        user.is_superuser or
        user.role in ['superadmin', 'administrador', 'bodeguero'] or
        (hasattr(user, 'can_view_inventory') and user.can_view_inventory()) or
        (hasattr(user, 'can_adjust_inventory') and user.can_adjust_inventory())
    )
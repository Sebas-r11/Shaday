from django.shortcuts import render, get_object_or_404, redirect
from django.views.generic import ListView, CreateView, DetailView, UpdateView, DeleteView, TemplateView, View
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.urls import reverse_lazy
from django.utils import timezone
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from django.db import models, transaction
from django.db.models import Sum, Q, Count
from django.http import HttpResponse, JsonResponse
from django.contrib import messages
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
from decimal import Decimal
from .models import Producto, Categoria, Subcategoria, Stock, Bodega, MovimientoInventario, Proveedor, ProductoProveedor, PresentacionProveedorProducto, OrdenCompraStock, ItemOrdenCompraStock, RecomendacionReposicion
from .forms import ProductoFilterForm, ProductoForm, ProveedorForm, ProductoProveedorForm, ProductoProveedorFormSet, BodegaForm

class InventarioRequiredMixin(UserPassesTestMixin):
    """Mixin para verificar permisos de inventario"""
    def test_func(self):
        return self.request.user.can_adjust_inventory()

class InventarioViewMixin(UserPassesTestMixin):
    """Mixin para permitir ver inventario (incluye bodega)"""
    def test_func(self):
        return (self.request.user.can_adjust_inventory() or 
                self.request.user.can_view_inventory())

class AdminOnlyMixin(UserPassesTestMixin):
    """Mixin para funciones que solo puede usar el administrador"""
    def test_func(self):
        return self.request.user.is_superuser or self.request.user.groups.filter(name='admin').exists()

class AdminInventarioMixin(UserPassesTestMixin):
    """Mixin para funciones estratégicas de inventario (solo administradores)"""
    def test_func(self):
        return (self.request.user.is_superuser or 
                self.request.user.role in ['superadmin', 'administrador'])

class ProductoListView(InventarioViewMixin, ListView):
    model = Producto
    template_name = 'inventario/producto_list.html'
    context_object_name = 'productos'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = Producto.objects.select_related('categoria', 'subcategoria').prefetch_related('stock__bodega')
        
        # Aplicar filtros
        form = ProductoFilterForm(self.request.GET)
        if form.is_valid():
            # Filtro por texto
            search = form.cleaned_data.get('search')
            if search:
                queryset = queryset.filter(
                    Q(codigo__icontains=search) |
                    Q(nombre__icontains=search)
                )
            
            # Filtro por categoría
            categoria = form.cleaned_data.get('categoria')
            if categoria:
                queryset = queryset.filter(categoria=categoria)
            
            # Filtro por subcategoría
            subcategoria = form.cleaned_data.get('subcategoria')
            if subcategoria:
                queryset = queryset.filter(subcategoria=subcategoria)
            
            # Filtro por rango de precios
            precio_min = form.cleaned_data.get('precio_min')
            if precio_min:
                queryset = queryset.filter(precio_minorista__gte=precio_min)
            
            precio_max = form.cleaned_data.get('precio_max')
            if precio_max:
                queryset = queryset.filter(precio_minorista__lte=precio_max)
            
            # Filtro por estado activo
            activo = form.cleaned_data.get('activo')
            if activo:
                queryset = queryset.filter(activo=(activo == 'True'))
            
            # Filtro por stock
            stock_status = form.cleaned_data.get('stock_status')
            if stock_status:
                if stock_status == 'sin_stock':
                    # Productos sin stock
                    productos_sin_stock = []
                    for producto in queryset:
                        stock_total = producto.stock.aggregate(total=Sum('cantidad'))['total'] or 0
                        if stock_total == 0:
                            productos_sin_stock.append(producto.id)
                    queryset = queryset.filter(id__in=productos_sin_stock)
                
                elif stock_status == 'bajo_minimo':
                    # Productos bajo el mínimo
                    productos_bajo_minimo = []
                    for producto in queryset:
                        stock_total = producto.stock.aggregate(total=Sum('cantidad'))['total'] or 0
                        if stock_total <= producto.stock_minimo:
                            productos_bajo_minimo.append(producto.id)
                    queryset = queryset.filter(id__in=productos_bajo_minimo)
                
                elif stock_status == 'disponible':
                    # Productos con stock disponible
                    productos_con_stock = []
                    for producto in queryset:
                        stock_total = producto.stock.aggregate(total=Sum('cantidad'))['total'] or 0
                        if stock_total > 0:
                            productos_con_stock.append(producto.id)
                    queryset = queryset.filter(id__in=productos_con_stock)
        
        return queryset.order_by('codigo')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter_form'] = ProductoFilterForm(self.request.GET)
        return context

class ProductoCreateView(AdminOnlyMixin, CreateView):
    model = Producto
    form_class = ProductoForm
    template_name = 'inventario/producto_form.html'
    success_url = reverse_lazy('inventario:producto_list')
    
    def form_valid(self, form):
        form.instance.usuario_creacion = self.request.user
        
        # Calcular precios basados en porcentajes si se proporcionaron
        costo = form.cleaned_data.get('costo_promedio', Decimal('0'))
        margen_min = form.cleaned_data.get('margen_minorista')
        margen_may = form.cleaned_data.get('margen_mayorista')
        
        if costo and margen_min is not None:
            precio_minorista = costo * (1 + (margen_min / Decimal('100')))
            form.instance.precio_minorista = precio_minorista.quantize(Decimal('0.01'))
        
        if costo and margen_may is not None:
            precio_mayorista = costo * (1 + (margen_may / Decimal('100')))
            form.instance.precio_mayorista = precio_mayorista.quantize(Decimal('0.01'))
        
        return super().form_valid(form)

class ProductoDetailView(InventarioViewMixin, DetailView):
    model = Producto
    template_name = 'inventario/producto_detail.html'
    context_object_name = 'producto'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Obtener stock por bodegas
        context['stocks'] = Stock.objects.filter(
            producto=self.object
        ).select_related('bodega', 'variante')
        return context

class ProductoUpdateView(AdminOnlyMixin, UpdateView):
    model = Producto
    form_class = ProductoForm
    template_name = 'inventario/producto_form.html'
    success_url = reverse_lazy('inventario:producto_list')
    
    def form_valid(self, form):
        # Calcular precios basados en porcentajes si se proporcionaron
        costo = form.cleaned_data.get('costo_promedio', Decimal('0'))
        margen_min = form.cleaned_data.get('margen_minorista')
        margen_may = form.cleaned_data.get('margen_mayorista')
        
        if costo and margen_min is not None:
            precio_minorista = costo * (1 + (margen_min / Decimal('100')))
            form.instance.precio_minorista = precio_minorista.quantize(Decimal('0.01'))
        
        if costo and margen_may is not None:
            precio_mayorista = costo * (1 + (margen_may / Decimal('100')))
            form.instance.precio_mayorista = precio_mayorista.quantize(Decimal('0.01'))
        
        return super().form_valid(form)

class CategoriaListView(InventarioViewMixin, ListView):
    model = Categoria
    template_name = 'inventario/categoria_list.html'
    context_object_name = 'categorias'

class CategoriaCreateView(AdminOnlyMixin, CreateView):
    model = Categoria
    template_name = 'inventario/categoria_form.html'
    fields = ['nombre', 'descripcion']
    success_url = reverse_lazy('inventario:categoria_list')

class CategoriaUpdateView(AdminOnlyMixin, UpdateView):
    model = Categoria
    template_name = 'inventario/categoria_form.html'
    fields = ['nombre', 'descripcion']
    success_url = reverse_lazy('inventario:categoria_list')

def exportar_productos_excel(request):
    """Vista para exportar productos a Excel"""
    if not request.user.is_authenticated:
        return HttpResponse("No autorizado", status=401)
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"
    
    # Configurar estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Definir headers basándose en permisos del usuario
    if request.user.can_see_costs():
        headers = [
            'Código', 'Nombre', 'Categoría', 'Subcategoría', 
            'Costo Promedio', 'Precio Minorista', 'Precio Mayorista',
            'Stock Total', 'Stock Mínimo', 'Estado', 'Fecha Creación'
        ]
    else:
        headers = [
            'Código', 'Nombre', 'Categoría', 'Subcategoría', 
            'Precio Minorista', 'Precio Mayorista',
            'Stock Total', 'Stock Mínimo', 'Estado', 'Fecha Creación'
        ]
    
    # Escribir headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Obtener productos con los mismos filtros que la vista de lista
    productos_view = ProductoListView()
    productos_view.request = request
    productos = productos_view.get_queryset()
    
    # Escribir datos
    row_num = 2
    for producto in productos:
        # Calcular stock total
        stock_total = producto.stock.aggregate(total=Sum('cantidad'))['total'] or 0
        
        if request.user.can_see_costs():
            row_data = [
                producto.codigo,
                producto.nombre,
                producto.categoria.nombre,
                producto.subcategoria.nombre,
                float(producto.costo_promedio),
                float(producto.precio_minorista),
                float(producto.precio_mayorista),
                stock_total,
                producto.stock_minimo,
                'Activo' if producto.activo else 'Inactivo',
                producto.fecha_creacion.strftime('%d/%m/%Y %H:%M')
            ]
        else:
            row_data = [
                producto.codigo,
                producto.nombre,
                producto.categoria.nombre,
                producto.subcategoria.nombre,
                float(producto.precio_minorista),
                float(producto.precio_mayorista),
                stock_total,
                producto.stock_minimo,
                'Activo' if producto.activo else 'Inactivo',
                producto.fecha_creacion.strftime('%d/%m/%Y %H:%M')
            ]
        
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)
        
        row_num += 1
    
    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'productos_reyes_{timestamp}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    # Guardar workbook en la respuesta
    wb.save(response)
    
    return response

# Subcategorías
class SubcategoriaListView(InventarioViewMixin, ListView):
    model = Subcategoria
    template_name = 'inventario/subcategoria_list.html'
    context_object_name = 'subcategorias'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = Subcategoria.objects.select_related('categoria').all()
        
        # Filtro por nombre de subcategoría
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(nombre__icontains=search)
        
        # Filtro por categoría padre
        categoria_id = self.request.GET.get('categoria')
        if categoria_id:
            try:
                queryset = queryset.filter(categoria_id=categoria_id)
            except (ValueError, TypeError):
                pass
        
        return queryset.order_by('categoria__nombre', 'nombre')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['categorias'] = Categoria.objects.all().order_by('nombre')
        context['search_query'] = self.request.GET.get('search', '')
        context['selected_categoria'] = self.request.GET.get('categoria', '')
        
        # Contar subcategorías por categoría para mostrar en filtros
        categorias_con_count = Categoria.objects.annotate(
            subcategorias_count=models.Count('subcategorias')
        ).order_by('nombre')
        context['categorias_con_count'] = categorias_con_count
        
        return context

class SubcategoriaCreateView(AdminOnlyMixin, CreateView):
    model = Subcategoria
    template_name = 'inventario/subcategoria_form.html'
    fields = ['nombre', 'descripcion', 'categoria']
    success_url = reverse_lazy('inventario:subcategoria_list')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['categorias'] = Categoria.objects.all()
        return context

class SubcategoriaUpdateView(AdminOnlyMixin, UpdateView):
    model = Subcategoria
    template_name = 'inventario/subcategoria_form.html'
    fields = ['nombre', 'descripcion', 'categoria']
    success_url = reverse_lazy('inventario:subcategoria_list')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['categorias'] = Categoria.objects.all()
        return context

class SubcategoriaDeleteView(AdminOnlyMixin, DeleteView):
    model = Subcategoria
    template_name = 'inventario/subcategoria_confirm_delete.html'
    success_url = reverse_lazy('inventario:subcategoria_list')

# Bodegas
class BodegaListView(InventarioViewMixin, ListView):
    model = Bodega
    template_name = 'inventario/bodega_list.html'
    context_object_name = 'bodegas'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Agregar estadísticas por bodega
        for bodega in context['bodegas']:
            bodega.total_productos = Stock.objects.filter(bodega=bodega).count()
            bodega.stock_total = Stock.objects.filter(bodega=bodega).aggregate(
                total=Sum('cantidad')
            )['total'] or 0
        return context

class BodegaCreateView(AdminOnlyMixin, CreateView):
    model = Bodega
    form_class = BodegaForm
    template_name = 'inventario/bodega_form.html'
    success_url = reverse_lazy('inventario:bodega_list')
    
    def form_valid(self, form):
        messages.success(self.request, f'✅ Bodega "{form.cleaned_data["nombre"]}" creada exitosamente.')
        return super().form_valid(form)

class BodegaDetailView(InventarioViewMixin, DetailView):
    model = Bodega
    template_name = 'inventario/bodega_detail.html'
    context_object_name = 'bodega'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Stock por productos en esta bodega
        context['stocks'] = Stock.objects.filter(
            bodega=self.object
        ).select_related('producto', 'variante').order_by('producto__nombre')
        
        # Movimientos recientes en esta bodega
        context['movimientos_recientes'] = MovimientoInventario.objects.filter(
            bodega=self.object
        ).select_related('producto', 'usuario').order_by('-fecha_movimiento')[:10]
        
        return context

class BodegaUpdateView(AdminOnlyMixin, UpdateView):
    model = Bodega
    form_class = BodegaForm
    template_name = 'inventario/bodega_form.html'
    success_url = reverse_lazy('inventario:bodega_list')
    
    def form_valid(self, form):
        messages.success(self.request, f'✅ Bodega "{form.cleaned_data["nombre"]}" actualizada exitosamente.')
        return super().form_valid(form)

class BodegaDeleteView(AdminOnlyMixin, DeleteView):
    model = Bodega
    template_name = 'inventario/bodega_confirm_delete.html'
    success_url = reverse_lazy('inventario:bodega_list')
    
    def get_context_data(self, **kwargs):
        """Añadir información para la confirmación"""
        context = super().get_context_data(**kwargs)
        bodega = self.get_object()
        
        # Información para la validación
        context['stock_actual'] = Stock.objects.filter(bodega=bodega, cantidad__gt=0).count()
        context['total_movimientos'] = MovimientoInventario.objects.filter(bodega=bodega).count()
        context['stock_historico'] = Stock.objects.filter(bodega=bodega).count()  # Incluye registros con cantidad 0
        
        # Últimos movimientos para contexto
        context['ultimos_movimientos'] = MovimientoInventario.objects.filter(
            bodega=bodega
        ).select_related('producto', 'usuario').order_by('-fecha_movimiento')[:5]
        
        return context
    
    def post(self, request, *args, **kwargs):
        """Verificación adicional antes de procesar el formulario"""
        self.object = self.get_object()
        
        # Verificación de seguridad adicional
        stock_count = Stock.objects.filter(bodega=self.object, cantidad__gt=0).count()
        if stock_count > 0:
            messages.error(
                request, 
                f'🚫 ACCIÓN BLOQUEADA: La bodega "{self.object.nombre}" tiene {stock_count} productos con stock. '
                f'No se puede proceder con la eliminación.'
            )
            return redirect('inventario:bodega_detail', pk=self.object.pk)
        
        # Si pasa la verificación, proceder con el método normal
        return super().post(request, *args, **kwargs)
    
    def delete(self, request, *args, **kwargs):
        """Validar antes de eliminar la bodega con lógica mejorada"""
        self.object = self.get_object()
        
        # 1. Verificar si tiene stock actual (doble verificación)
        stock_count = Stock.objects.filter(bodega=self.object, cantidad__gt=0).count()
        if stock_count > 0:
            messages.error(
                request, 
                f'❌ VALIDACIÓN FALLÓ: La bodega "{self.object.nombre}" tiene {stock_count} productos con stock actual. '
                f'Esta acción no debería haber sido posible.'
            )
            return redirect('inventario:bodega_detail', pk=self.object.pk)
        
        # 2. Verificar si tiene registros de stock (incluso con cantidad 0)
        stock_historico_count = Stock.objects.filter(bodega=self.object).count()
        if stock_historico_count > 0:
            messages.warning(
                request, 
                f'⚠️ La bodega "{self.object.nombre}" tiene registros de inventario ({stock_historico_count} productos). Se marcará como inactiva para preservar el historial.'
            )
            self.object.activa = False
            self.object.save()
            messages.success(request, f'✅ Bodega "{self.object.nombre}" marcada como inactiva.')
            return redirect('inventario:bodega_list')
        
        # 3. Verificar si tiene movimientos de inventario
        movimientos_count = MovimientoInventario.objects.filter(bodega=self.object).count()
        if movimientos_count > 0:
            messages.warning(
                request, 
                f'⚠️ La bodega "{self.object.nombre}" tiene {movimientos_count} movimientos registrados. Se marcará como inactiva para preservar el historial.'
            )
            self.object.activa = False
            self.object.save()
            messages.success(request, f'✅ Bodega "{self.object.nombre}" marcada como inactiva.')
            return redirect('inventario:bodega_list')
        
        # 4. Solo permitir eliminación completa si es una bodega completamente nueva
        messages.success(request, f'✅ Bodega "{self.object.nombre}" eliminada exitosamente.')
        return super().delete(request, *args, **kwargs)

# Stock
class StockListView(InventarioViewMixin, ListView):
    model = Stock
    template_name = 'inventario/stock_list.html'
    context_object_name = 'stocks'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = Stock.objects.select_related('producto', 'bodega', 'variante').all()
        
        # Filtros
        producto_search = self.request.GET.get('producto')
        if producto_search:
            queryset = queryset.filter(
                Q(producto__codigo__icontains=producto_search) |
                Q(producto__nombre__icontains=producto_search)
            )
        
        bodega_id = self.request.GET.get('bodega')
        if bodega_id:
            try:
                queryset = queryset.filter(bodega_id=bodega_id)
            except (ValueError, TypeError):
                pass
        
        stock_status = self.request.GET.get('stock_status')
        if stock_status == 'sin_stock':
            queryset = queryset.filter(cantidad=0)
        elif stock_status == 'bajo_minimo':
            queryset = queryset.filter(cantidad__lte=models.F('producto__stock_minimo'))
        
        return queryset.order_by('bodega__nombre', 'producto__nombre')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['bodegas'] = Bodega.objects.filter(activa=True)
        context['filtros'] = {
            'producto': self.request.GET.get('producto', ''),
            'bodega': self.request.GET.get('bodega', ''),
            'stock_status': self.request.GET.get('stock_status', ''),
        }
        return context

class StockDetailView(InventarioViewMixin, DetailView):
    model = Stock
    template_name = 'inventario/stock_detail.html'
    context_object_name = 'stock'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Movimientos relacionados con este stock
        context['movimientos'] = MovimientoInventario.objects.filter(
            producto=self.object.producto,
            bodega=self.object.bodega
        ).select_related('usuario').order_by('-fecha_movimiento')[:20]
        return context

# Movimientos de inventario
class MovimientoInventarioListView(AdminInventarioMixin, ListView):
    """Vista de movimientos de inventario - Solo para administradores"""
    model = MovimientoInventario
    template_name = 'inventario/movimiento_list.html'
    context_object_name = 'movimientos'
    paginate_by = 30
    
    def get_queryset(self):
        queryset = MovimientoInventario.objects.select_related(
            'producto', 'bodega', 'usuario', 'variante'
        ).all()
        
        # Filtros
        producto_search = self.request.GET.get('producto')
        if producto_search:
            queryset = queryset.filter(
                Q(producto__codigo__icontains=producto_search) |
                Q(producto__nombre__icontains=producto_search)
            )
        
        bodega_id = self.request.GET.get('bodega')
        if bodega_id:
            try:
                queryset = queryset.filter(bodega_id=bodega_id)
            except (ValueError, TypeError):
                pass
        
        tipo_movimiento = self.request.GET.get('tipo')
        if tipo_movimiento:
            queryset = queryset.filter(tipo_movimiento=tipo_movimiento)
        
        fecha_desde = self.request.GET.get('fecha_desde')
        if fecha_desde:
            try:
                from datetime import datetime
                fecha = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
                queryset = queryset.filter(fecha_movimiento__date__gte=fecha)
            except ValueError:
                pass
        
        fecha_hasta = self.request.GET.get('fecha_hasta')
        if fecha_hasta:
            try:
                from datetime import datetime
                fecha = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
                queryset = queryset.filter(fecha_movimiento__date__lte=fecha)
            except ValueError:
                pass
        
        return queryset.order_by('-fecha_movimiento')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['bodegas'] = Bodega.objects.filter(activa=True)
        context['filtros'] = {
            'producto': self.request.GET.get('producto', ''),
            'bodega': self.request.GET.get('bodega', ''),
            'tipo': self.request.GET.get('tipo', ''),
            'fecha_desde': self.request.GET.get('fecha_desde', ''),
            'fecha_hasta': self.request.GET.get('fecha_hasta', ''),
        }
        return context

# Transferencias entre bodegas
def transferencia_producto(request):
    """Vista para manejar transferencias entre bodegas - Solo para administradores"""
    # Verificar permisos - Los bodegueros no pueden hacer transferencias
    if not request.user.is_authenticated:
        messages.error(request, 'Debe iniciar sesión para acceder a esta función.')
        return redirect('accounts:login')
    
    if request.user.role == 'bodega':
        messages.error(request, 'Los bodegueros no tienen permisos para realizar transferencias entre bodegas.')
        return redirect('inventario:stock_list')
    
    if not request.user.role in ['superadmin', 'administrador']:
        messages.error(request, 'No tiene permisos para realizar transferencias entre bodegas.')
        return redirect('inventario:stock_list')
    
    if request.method == 'POST':
        try:
            producto_id = request.POST.get('producto_id')
            bodega_origen_id = request.POST.get('bodega_origen')
            bodega_destino_id = request.POST.get('bodega_destino')
            cantidad = int(request.POST.get('cantidad', 0))
            observaciones = request.POST.get('observaciones', '')
            
            # Validaciones
            if not all([producto_id, bodega_origen_id, bodega_destino_id, cantidad]):
                messages.error(request, 'Todos los campos son obligatorios.')
                return render(request, 'inventario/transferencia_form.html', {
                    'productos': Producto.objects.all(),
                    'bodegas': Bodega.objects.filter(activa=True)
                })
            
            if bodega_origen_id == bodega_destino_id:
                messages.error(request, 'La bodega de origen debe ser diferente a la de destino.')
                return render(request, 'inventario/transferencia_form.html', {
                    'productos': Producto.objects.all(),
                    'bodegas': Bodega.objects.filter(activa=True)
                })
            
            if cantidad <= 0:
                messages.error(request, 'La cantidad debe ser mayor a cero.')
                return render(request, 'inventario/transferencia_form.html', {
                    'productos': Producto.objects.all(),
                    'bodegas': Bodega.objects.filter(activa=True)
                })
            
            producto = get_object_or_404(Producto, id=producto_id)
            bodega_origen = get_object_or_404(Bodega, id=bodega_origen_id)
            bodega_destino = get_object_or_404(Bodega, id=bodega_destino_id)
            
            # Verificar stock disponible en bodega origen
            stock_origen = Stock.objects.filter(
                producto=producto,
                bodega=bodega_origen
            ).first()
            
            if not stock_origen or stock_origen.cantidad < cantidad:
                messages.error(request, f'Stock insuficiente en {bodega_origen.nombre}. Stock disponible: {stock_origen.cantidad if stock_origen else 0}')
                return render(request, 'inventario/transferencia_form.html', {
                    'productos': Producto.objects.all(),
                    'bodegas': Bodega.objects.filter(activa=True)
                })
            
            # Realizar transferencia en una transacción
            with transaction.atomic():
                # 1. Reducir stock en bodega origen
                stock_origen.cantidad -= cantidad
                stock_origen.save()
                
                # 2. Aumentar stock en bodega destino (crear si no existe)
                stock_destino, created = Stock.objects.get_or_create(
                    producto=producto,
                    bodega=bodega_destino,
                    defaults={'cantidad': 0, 'cantidad_reservada': 0}
                )
                stock_destino.cantidad += cantidad
                stock_destino.save()
                
                # 3. Registrar movimiento de salida en bodega origen
                movimiento_salida = MovimientoInventario.objects.create(
                    producto=producto,
                    bodega=bodega_origen,
                    bodega_destino=bodega_destino,
                    tipo_movimiento='transferencia',
                    motivo='transferencia',
                    cantidad=cantidad,
                    costo_unitario=producto.costo_promedio,
                    observaciones=f'Transferencia a {bodega_destino.nombre}. {observaciones}',
                    usuario=request.user
                )
                
                # 4. Registrar movimiento de entrada en bodega destino
                MovimientoInventario.objects.create(
                    producto=producto,
                    bodega=bodega_destino,
                    tipo_movimiento='entrada',
                    motivo='transferencia',
                    cantidad=cantidad,
                    costo_unitario=producto.costo_promedio,
                    observaciones=f'Transferencia desde {bodega_origen.nombre}. {observaciones}',
                    usuario=request.user
                )
            
            # Verificar si se solicita generar PDF directamente
            if request.POST.get('generar_pdf'):
                return generar_pdf_transferencia(request, movimiento_salida.id)
            
            # Mensaje con opción de imprimir
            messages.success(request, f'Transferencia realizada exitosamente: {cantidad} unidades de {producto.nombre} de {bodega_origen.nombre} a {bodega_destino.nombre}')
            
            # Redirigir con parámetro para mostrar opción de impresión
            return redirect(f'{reverse_lazy("inventario:stock_list")}?transferencia_id={movimiento_salida.id}&success=1')
            
        except Exception as e:
            messages.error(request, f'Error al realizar la transferencia: {str(e)}')
            return render(request, 'inventario/transferencia_form.html', {
                'productos': Producto.objects.all(),
                'bodegas': Bodega.objects.filter(activa=True)
            })
    
    # GET request - mostrar formulario
    context = {
        'productos': Producto.objects.filter(activo=True).order_by('nombre'),
        'bodegas': Bodega.objects.filter(activa=True).order_by('nombre'),
        'stocks': Stock.objects.select_related('producto', 'bodega').filter(cantidad__gt=0).order_by('bodega__nombre', 'producto__nombre')
    }
    
    return render(request, 'inventario/transferencia_form.html', context)

def generar_pdf_transferencia(request, movimiento_id):
    """Vista para generar PDF del documento de transferencia"""
    movimiento = get_object_or_404(MovimientoInventario, id=movimiento_id, tipo_movimiento='transferencia')
    
    # Crear respuesta PDF
    response = HttpResponse(content_type='application/pdf')
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'transferencia_{movimiento.producto.codigo}_{timestamp}.pdf'
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    # Crear PDF con reportlab
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.graphics.shapes import Drawing, Line
    
    # Configurar documento
    doc = SimpleDocTemplate(response, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
    styles = getSampleStyleSheet()
    story = []
    
    # Estilo personalizado para título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1,  # Center
        textColor=colors.navy
    )
    
    # Estilo para subtítulos
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Heading2'],
        fontSize=14,
        spaceAfter=12,
        textColor=colors.darkblue
    )
    
    # Título del documento
    story.append(Paragraph("DOCUMENTO DE TRANSFERENCIA DE INVENTARIO", title_style))
    story.append(Paragraph("DistribucioneShaddai", styles['Normal']))
    story.append(Spacer(1, 20))
    
    # Información de la transferencia
    story.append(Paragraph("INFORMACIÓN GENERAL", subtitle_style))
    
    transfer_data = [
        ['Fecha de Transferencia:', movimiento.fecha_movimiento.strftime('%d/%m/%Y %H:%M')],
        ['N° de Documento:', str(movimiento.id)[:8].upper()],
        ['Usuario Responsable:', movimiento.usuario.get_full_name() or movimiento.usuario.username],
        ['', ''],  # Línea vacía
    ]
    
    transfer_table = Table(transfer_data, colWidths=[2*inch, 3*inch])
    transfer_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -2), 1, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    
    story.append(transfer_table)
    story.append(Spacer(1, 20))
    
    # Información de bodegas
    story.append(Paragraph("MOVIMIENTO ENTRE BODEGAS", subtitle_style))
    
    bodega_data = [
        ['BODEGA ORIGEN', 'BODEGA DESTINO'],
        [movimiento.bodega.nombre, movimiento.bodega_destino.nombre if movimiento.bodega_destino else 'N/A'],
        [movimiento.bodega.direccion or 'Sin dirección', movimiento.bodega_destino.direccion or 'Sin dirección' if movimiento.bodega_destino else ''],
    ]
    
    bodega_table = Table(bodega_data, colWidths=[2.5*inch, 2.5*inch])
    bodega_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    story.append(bodega_table)
    story.append(Spacer(1, 20))
    
    # Detalle del producto
    story.append(Paragraph("DETALLE DEL PRODUCTO", subtitle_style))
    
    producto_data = [
        ['Código', 'Descripción', 'Cantidad', 'Costo Unit.', 'Total'],
        [
            movimiento.producto.codigo,
            movimiento.producto.nombre,
            str(movimiento.cantidad),
            f"${movimiento.costo_unitario:,.0f}",
            f"${movimiento.costo_unitario * movimiento.cantidad:,.0f}"
        ]
    ]
    
    producto_table = Table(producto_data, colWidths=[1*inch, 2.5*inch, 0.8*inch, 1*inch, 1*inch])
    producto_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    story.append(producto_table)
    story.append(Spacer(1, 20))
    
    # Observaciones
    if movimiento.observaciones:
        story.append(Paragraph("OBSERVACIONES", subtitle_style))
        story.append(Paragraph(movimiento.observaciones, styles['Normal']))
        story.append(Spacer(1, 30))
    else:
        story.append(Spacer(1, 50))
    
    # Firmas
    story.append(Paragraph("FIRMAS Y AUTORIZACIONES", subtitle_style))
    story.append(Spacer(1, 40))
    
    firma_data = [
        ['_____________________', '_____________________', '_____________________'],
        ['ENTREGA', 'RECIBE', 'AUTORIZA'],
        ['Responsable Bodega Origen', 'Responsable Bodega Destino', 'Supervisor'],
    ]
    
    firma_table = Table(firma_data, colWidths=[2*inch, 2*inch, 2*inch])
    firma_table.setStyle(TableStyle([
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 1), (-1, 1), 10),
        ('FONTNAME', (0, 2), (-1, 2), 'Helvetica'),
        ('FONTSIZE', (0, 2), (-1, 2), 8),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    story.append(firma_table)
    
    # Footer
    story.append(Spacer(1, 30))
    story.append(Paragraph(f"Documento generado automáticamente el {datetime.now().strftime('%d/%m/%Y %H:%M')}", 
                          ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, alignment=1)))
    
    # Construir PDF
    doc.build(story)
    
    return response

# Vista mejorada para productos con proveedores
class ProductoDetailViewConProveedores(InventarioViewMixin, DetailView):
    model = Producto
    template_name = 'inventario/producto_detail_proveedores.html'
    context_object_name = 'producto'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Stock por bodegas
        context['stocks'] = Stock.objects.filter(
            producto=self.object
        ).select_related('bodega', 'variante')
        
        # Proveedores activos para este producto
        context['producto_proveedores'] = ProductoProveedor.objects.filter(
            producto=self.object,
            activo=True
        ).select_related('proveedor').order_by('proveedor_preferido', 'precio_compra')
        
        return context

# Vista para gestionar proveedores de un producto
class ProductoProveedoresView(AdminOnlyMixin, UpdateView):
    model = Producto
    template_name = 'inventario/producto_proveedores_form.html'
    fields = []  # No campos del producto principal
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        if self.request.POST:
            context['formset'] = ProductoProveedorFormSet(
                self.request.POST, 
                instance=self.object
            )
        else:
            context['formset'] = ProductoProveedorFormSet(instance=self.object)
        
        return context
    
    def form_valid(self, form):
        context = self.get_context_data()
        formset = context['formset']
        
        if formset.is_valid():
            with transaction.atomic():
                self.object = form.save()
                formset.instance = self.object
                formset.save()
                
                # Actualizar costo promedio basado en proveedores
                self.object.actualizar_costo_promedio()
                
            messages.success(self.request, f'Proveedores del producto {self.object.codigo} actualizados exitosamente.')
            return redirect('inventario:producto_detail', pk=self.object.pk)
        else:
            return self.render_to_response(self.get_context_data(form=form))

# Vistas de Alertas de Stock y Órdenes de Compra
class AlertasStockView(AdminInventarioMixin, ListView):
    """Vista de productos con stock bajo o crítico - Solo para administradores"""
    model = Producto
    template_name = 'inventario/alertas_stock.html'
    context_object_name = 'productos'
    paginate_by = 50
    
    def get_queryset(self):
        # Usar el método que agregamos al modelo
        productos = Producto.productos_con_alerta_stock()
        
        # Filtros adicionales
        nivel = self.request.GET.get('nivel')
        if nivel == 'critico':
            productos = [p for p in productos if p.stock_critico]
        elif nivel == 'bajo':
            productos = [p for p in productos if p.stock_cerca_minimo]
        
        # Filtro por categoría
        categoria_id = self.request.GET.get('categoria')
        if categoria_id:
            productos = productos.filter(categoria_id=categoria_id)
        
        return productos
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Estadísticas de alertas
        todos_productos = Producto.productos_con_alerta_stock()
        context['total_alertas'] = todos_productos.count()
        context['productos_criticos'] = len([p for p in todos_productos if p.stock_critico])
        context['productos_bajo_stock'] = len([p for p in todos_productos if p.stock_cerca_minimo])
        
        # Categorías para filtro
        context['categorias'] = Categoria.objects.filter(activa=True)
        
        # Filtros aplicados
        context['nivel_filtro'] = self.request.GET.get('nivel', '')
        context['categoria_filtro'] = self.request.GET.get('categoria', '')
        
        return context

class GenerarOrdenCompraView(AdminOnlyMixin, TemplateView):
    """Vista para generar órdenes de compra desde alertas de stock"""
    template_name = 'inventario/generar_orden_compra.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Obtener productos seleccionados
        productos_ids = self.request.GET.getlist('productos')
        if productos_ids:
            productos = Producto.objects.filter(id__in=productos_ids).prefetch_related(
                'producto_proveedores__proveedor'
            )
            context['productos_seleccionados'] = productos
            
            # Agrupar por proveedor preferido
            ordenes_por_proveedor = {}
            productos_sin_proveedor = []
            
            for producto in productos:
                proveedor_preferido = producto.proveedor_preferido
                if proveedor_preferido:
                    if proveedor_preferido.proveedor not in ordenes_por_proveedor:
                        ordenes_por_proveedor[proveedor_preferido.proveedor] = []
                    ordenes_por_proveedor[proveedor_preferido.proveedor].append({
                        'producto': producto,
                        'producto_proveedor': proveedor_preferido,
                        'cantidad_sugerida': producto.cantidad_sugerida_compra
                    })
                else:
                    # Buscar el mejor proveedor disponible
                    mejor_proveedor = producto.mejor_precio_proveedor
                    if mejor_proveedor:
                        if mejor_proveedor.proveedor not in ordenes_por_proveedor:
                            ordenes_por_proveedor[mejor_proveedor.proveedor] = []
                        ordenes_por_proveedor[mejor_proveedor.proveedor].append({
                            'producto': producto,
                            'producto_proveedor': mejor_proveedor,
                            'cantidad_sugerida': producto.cantidad_sugerida_compra
                        })
                    else:
                        productos_sin_proveedor.append(producto)
            
            context['ordenes_por_proveedor'] = ordenes_por_proveedor
            context['productos_sin_proveedor'] = productos_sin_proveedor
        
        return context
    
    def post(self, request, *args, **kwargs):
        """Crear las órdenes de compra con selección manual"""
        try:
            with transaction.atomic():
                # Obtener productos seleccionados
                productos_seleccionados = []
                for key, value in request.POST.items():
                    if key == 'productos':
                        productos_seleccionados.extend(request.POST.getlist('productos'))
                
                if not productos_seleccionados:
                    messages.error(request, 'No se seleccionaron productos')
                    return redirect('inventario:alertas_stock')
                
                # Agrupar por proveedor seleccionado
                ordenes_por_proveedor = {}
                
                for producto_id in productos_seleccionados:
                    # Obtener proveedor, cantidad, precio y presentación seleccionados
                    proveedor_key = f'proveedor_{producto_id}'
                    cantidad_key = f'cantidad_{producto_id}'
                    precio_key = f'precio_{producto_id}'
                    presentacion_key = f'presentacion_{producto_id}'
                    
                    if proveedor_key not in request.POST or cantidad_key not in request.POST:
                        continue
                    
                    producto_proveedor_id = request.POST[proveedor_key]
                    cantidad = int(request.POST[cantidad_key])
                    precio_personalizado = None
                    presentacion_proveedor_id = request.POST.get(presentacion_key)
                    
                    # Obtener precio personalizado si está disponible
                    if precio_key in request.POST and request.POST[precio_key]:
                        try:
                            precio_personalizado = Decimal(str(request.POST[precio_key]))
                        except (ValueError, TypeError):
                            precio_personalizado = None
                    
                    if not producto_proveedor_id or cantidad <= 0:
                        continue
                    
                    # Obtener objetos
                    producto = Producto.objects.get(id=producto_id)
                    producto_proveedor = ProductoProveedor.objects.select_related('proveedor').get(id=producto_proveedor_id)
                    
                    # Obtener presentación si fue seleccionada
                    presentacion_proveedor = None
                    if presentacion_proveedor_id:
                        try:
                            presentacion_proveedor = PresentacionProveedorProducto.objects.get(id=presentacion_proveedor_id)
                        except PresentacionProveedorProducto.DoesNotExist:
                            pass
                    
                    # Agrupar por proveedor
                    proveedor = producto_proveedor.proveedor
                    if proveedor not in ordenes_por_proveedor:
                        ordenes_por_proveedor[proveedor] = []
                    
                    ordenes_por_proveedor[proveedor].append({
                        'producto': producto,
                        'producto_proveedor': producto_proveedor,
                        'presentacion_proveedor': presentacion_proveedor,
                        'cantidad': cantidad,
                        'precio_personalizado': precio_personalizado
                    })
                
                # Crear órdenes de compra por proveedor
                ordenes_creadas = []
                
                for proveedor, items in ordenes_por_proveedor.items():
                    # Crear la orden de compra
                    orden = OrdenCompraStock.objects.create(
                        proveedor=proveedor,
                        usuario_creacion=request.user,
                        generada_por_alerta=True,
                        observaciones=f"Orden generada por alertas de stock - {len(items)} productos seleccionados"
                    )
                    
                    # Crear los items
                    for item_info in items:
                        # Usar precio personalizado si está disponible, sino el del proveedor
                        precio_final = item_info['precio_personalizado'] if item_info['precio_personalizado'] is not None else item_info['producto_proveedor'].precio_compra
                        
                        # Si hay presentación específica del proveedor, usar su precio
                        if item_info['presentacion_proveedor']:
                            precio_final = item_info['presentacion_proveedor'].precio_compra_presentacion
                        
                        ItemOrdenCompraStock.objects.create(
                            orden_compra=orden,
                            producto=item_info['producto'],
                            producto_proveedor=item_info['producto_proveedor'],
                            presentacion_proveedor=item_info['presentacion_proveedor'],
                            cantidad_solicitada=item_info['cantidad'],
                            precio_unitario=precio_final,
                            descuento_porcentaje=item_info['producto_proveedor'].descuento_volumen if item_info['cantidad'] >= item_info['producto_proveedor'].cantidad_descuento else Decimal('0.00'),
                            observaciones=f"Stock actual: {item_info['producto'].stock_total}, Mínimo: {item_info['producto'].stock_minimo}. {'Presentación: ' + item_info['presentacion_proveedor'].presentacion_base.nombre + '. ' if item_info['presentacion_proveedor'] else ''}Precio {'personalizado' if item_info['precio_personalizado'] else 'del proveedor'}"
                        )
                    
                    # Calcular totales de la orden
                    orden.calcular_totales()
                    ordenes_creadas.append(orden)
                
                messages.success(
                    request, 
                    f"Se crearon {len(ordenes_creadas)} órdenes de compra exitosamente para {len(ordenes_por_proveedor)} proveedores."
                )
            
            return redirect('inventario:alertas_stock')
            
        except Exception as e:
            messages.error(request, f"Error al crear las órdenes: {str(e)}")
            return self.get(request, *args, **kwargs)

class OrdenCompraStockListView(AdminOnlyMixin, ListView):
    """Lista de órdenes de compra por alertas de stock"""
    model = OrdenCompraStock
    template_name = 'inventario/orden_compra_stock_list.html'
    context_object_name = 'ordenes'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = OrdenCompraStock.objects.select_related('proveedor', 'usuario_creacion')
        
        # Filtros
        estado = self.request.GET.get('estado')
        if estado:
            queryset = queryset.filter(estado=estado)
        
        proveedor_id = self.request.GET.get('proveedor')
        if proveedor_id:
            queryset = queryset.filter(proveedor_id=proveedor_id)
        
        return queryset.order_by('-fecha_creacion')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['proveedores'] = Proveedor.objects.filter(activo=True)
        context['estados'] = OrdenCompraStock.ESTADO_CHOICES
        return context

class OrdenCompraStockDetailView(AdminOnlyMixin, DetailView):
    """Detalle de orden de compra por stock"""
    model = OrdenCompraStock
    template_name = 'inventario/orden_compra_stock_detail.html'
    context_object_name = 'orden'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['items'] = self.object.items.select_related('producto', 'producto_proveedor')
        return context


@method_decorator(csrf_exempt, name='dispatch')
class CambiarEstadoOrdenCompraView(AdminOnlyMixin, View):
    """Vista para cambiar el estado de una orden de compra"""
    
    def post(self, request, pk):
        try:
            orden = get_object_or_404(OrdenCompraStock, pk=pk)
            nuevo_estado = request.POST.get('nuevo_estado')
            
            if nuevo_estado not in ['enviada', 'recibida', 'cancelada']:
                messages.error(request, 'Estado no válido')
                return redirect('inventario:orden_compra_stock_detail', pk=pk)
            
            # Validar transición de estado
            if orden.estado == 'cancelada':
                messages.error(request, 'No se puede modificar una orden cancelada')
                return redirect('inventario:orden_compra_stock_detail', pk=pk)
            
            if orden.estado == 'recibida':
                messages.error(request, 'No se puede modificar una orden ya recibida')
                return redirect('inventario:orden_compra_stock_detail', pk=pk)
            
            if nuevo_estado == 'recibida' and orden.estado != 'enviada':
                messages.error(request, 'Solo se pueden recibir órdenes enviadas')
                return redirect('inventario:orden_compra_stock_detail', pk=pk)
            
            # Cambiar estado
            orden.estado = nuevo_estado
            orden.save()
            
            # Mensaje de confirmación
            estados_msg = {
                'enviada': 'enviada al proveedor',
                'recibida': 'recibida correctamente',
                'cancelada': 'cancelada'
            }
            
            messages.success(request, f'Orden de compra {estados_msg[nuevo_estado]} exitosamente')
            
            # Si se marca como recibida, actualizar stock (opcional)
            if nuevo_estado == 'recibida':
                # Aquí se podría agregar lógica para actualizar stock automáticamente
                pass
            
            return redirect('inventario:orden_compra_stock_detail', pk=pk)
            
        except Exception as e:
            messages.error(request, f'Error al cambiar estado: {str(e)}')
            return redirect('inventario:orden_compra_stock_detail', pk=pk)

def generar_pdf_orden_compra(request, orden_id):
    """Vista para generar PDF de la orden de compra sin valores monetarios"""
    from django.http import HttpResponse
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from datetime import datetime
    import io
    
    orden = get_object_or_404(OrdenCompraStock, pk=orden_id)
    
    # Crear respuesta PDF
    response = HttpResponse(content_type='application/pdf')
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'orden_compra_{orden.codigo}_{timestamp}.pdf'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Crear PDF con reportlab
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
    
    # Contenido del PDF
    elements = []
    styles = getSampleStyleSheet()
    
    # Estilo personalizado para el título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=1,  # Centro
        textColor=colors.darkblue
    )
    
    # Estilo para subtítulos
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Heading2'],
        fontSize=12,
        spaceBefore=20,
        spaceAfter=10,
        textColor=colors.darkblue
    )
    
    # Encabezado del documento
    elements.append(Paragraph("DISTRIBUCIONE SHADDAI", title_style))
    elements.append(Paragraph("ORDEN DE COMPRA", title_style))
    elements.append(Spacer(1, 20))
    
    # Información básica de la orden
    info_data = [
        ['Código de Orden:', orden.codigo],
        ['Estado:', orden.get_estado_display()],
        ['Fecha de Creación:', orden.fecha_creacion.strftime('%d/%m/%Y %H:%M')],
        ['Tipo:', 'Generada por Alerta de Stock' if orden.generada_por_alerta else 'Generada Manualmente'],
    ]
    
    if orden.fecha_entrega_esperada:
        info_data.append(['Fecha de Entrega Esperada:', orden.fecha_entrega_esperada.strftime('%d/%m/%Y')])
    
    if orden.usuario_creacion:
        info_data.append(['Creada por:', orden.usuario_creacion.get_full_name() or orden.usuario_creacion.username])
    
    info_table = Table(info_data, colWidths=[2*inch, 3*inch])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    elements.append(info_table)
    elements.append(Spacer(1, 20))
    
    # Información del proveedor
    elements.append(Paragraph("INFORMACIÓN DEL PROVEEDOR", subtitle_style))
    
    proveedor_data = [
        ['Nombre:', orden.proveedor.nombre],
        ['NIT:', orden.proveedor.nit],
        ['Contacto Principal:', orden.proveedor.contacto_principal or 'No especificado'],
        ['Teléfono:', orden.proveedor.telefono or 'No especificado'],
        ['Teléfono Contacto:', orden.proveedor.telefono_contacto or 'No especificado'],
        ['Email:', orden.proveedor.email or 'No especificado'],
        ['Dirección:', orden.proveedor.direccion or 'No especificado'],
        ['Ciudad:', orden.proveedor.ciudad or 'No especificado'],
    ]
    
    proveedor_table = Table(proveedor_data, colWidths=[2*inch, 4*inch])
    proveedor_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    elements.append(proveedor_table)
    elements.append(Spacer(1, 20))
    
    # Detalles de los productos (SIN VALORES MONETARIOS)
    elements.append(Paragraph("PRODUCTOS SOLICITADOS", subtitle_style))
    
    # Encabezados de la tabla de productos
    productos_data = [
        ['Código', 'Producto', 'Cantidad', 'Unidad']
    ]
    
    # Items de la orden
    items = orden.items.select_related('producto', 'presentacion', 'presentacion_proveedor__presentacion_base').all()
    for item in items:
        productos_data.append([
            item.producto.codigo,
            Paragraph(item.producto.nombre[:40] + '...' if len(item.producto.nombre) > 40 else item.producto.nombre, styles['Normal']),
            str(item.cantidad_solicitada),
            item.nombre_presentacion
        ])
    
    productos_table = Table(productos_data, colWidths=[1.2*inch, 3.5*inch, 1*inch, 1*inch])
    productos_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (1, 1), (1, -1), 'LEFT'),  # Alinear nombres de productos a la izquierda
        ('ALIGN', (4, 1), (4, -1), 'LEFT'),  # Alinear observaciones a la izquierda
    ]))
    
    elements.append(productos_table)
    elements.append(Spacer(1, 20))
    
    # Observaciones generales
    if orden.observaciones:
        elements.append(Paragraph("OBSERVACIONES", subtitle_style))
        elements.append(Paragraph(orden.observaciones, styles['Normal']))
        elements.append(Spacer(1, 20))
    
    # Notas internas (si las hay)
    if orden.notas_internas:
        elements.append(Paragraph("NOTAS INTERNAS", subtitle_style))
        elements.append(Paragraph(orden.notas_internas, styles['Normal']))
        elements.append(Spacer(1, 20))
    
    # Pie de página
    elements.append(Spacer(1, 30))
    elements.append(Paragraph("___________________________________", styles['Normal']))
    elements.append(Paragraph("Firma del Responsable", styles['Normal']))
    elements.append(Spacer(1, 10))
    elements.append(Paragraph(f"Documento generado el {datetime.now().strftime('%d/%m/%Y a las %H:%M')}", styles['Italic']))
    
    # Construir PDF
    doc.build(elements)
    
    # Obtener el valor del buffer y escribirlo a la respuesta
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    
    return response

# API Views
def subcategorias_api(request):
    """API para obtener subcategorías por categoría"""
    categoria_id = request.GET.get('categoria')
    
    if categoria_id:
        subcategorias = Subcategoria.objects.filter(
            categoria_id=categoria_id, 
            activa=True
        ).values('id', 'nombre')
        return JsonResponse(list(subcategorias), safe=False)
    
    return JsonResponse([], safe=False)

def stock_api(request):
    """API para obtener stock actual de un producto en una bodega"""
    producto_id = request.GET.get('producto')
    bodega_id = request.GET.get('bodega')
    
    if producto_id and bodega_id:
        try:
            stock = Stock.objects.get(producto_id=producto_id, bodega_id=bodega_id)
            return JsonResponse({'stock': stock.cantidad})
        except Stock.DoesNotExist:
            return JsonResponse({'stock': 0})
    
    return JsonResponse({'error': 'Parámetros inválidos'}, status=400)

# Ajustes de Inventario
class AjusteInventarioView(AdminInventarioMixin, TemplateView):
    """Vista para mostrar el formulario de ajustes de inventario"""
    template_name = 'inventario/ajuste_inventario.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['productos'] = Producto.objects.filter(activo=True).order_by('codigo')
        context['bodegas'] = Bodega.objects.filter(activa=True).order_by('nombre')
        return context

@csrf_exempt
def crear_ajuste_inventario(request):
    """Vista para crear ajustes de inventario (entradas y salidas)"""
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'No autenticado'}, status=401)
    
    # Verificar permisos de inventario
    if not (request.user.is_superuser or 
            request.user.role in ['superadmin', 'administrador'] or
            request.user.can_adjust_inventory()):
        return JsonResponse({'error': 'Sin permisos para ajustar inventario'}, status=403)
    
    if request.method == 'POST':
        try:
            # Obtener datos del formulario
            producto_id = request.POST.get('producto_id')
            bodega_id = request.POST.get('bodega_id')
            tipo_movimiento = request.POST.get('tipo_movimiento')  # 'entrada' o 'salida'
            cantidad = int(request.POST.get('cantidad', 0))
            motivo_form = request.POST.get('motivo', '')
            observaciones = request.POST.get('observaciones', '')
            
            # Mapear motivos del formulario a opciones del modelo
            motivos_map = {
                'Compra': 'compra',
                'Ajuste inicial': 'ajuste_inventario',
                'Corrección': 'ajuste_inventario',
                'Producto dañado': 'producto_dañado',
                'Producto vencido': 'producto_vencido',
                'Pérdida': 'ajuste_inventario',
                'Promoción': 'ajuste_inventario',
                'Otro': 'ajuste_inventario',
            }
            
            motivo = motivos_map.get(motivo_form, 'ajuste_inventario')
            
            # Validaciones
            if not all([producto_id, bodega_id, tipo_movimiento, cantidad]):
                return JsonResponse({'error': 'Todos los campos son requeridos'}, status=400)
            
            if cantidad <= 0:
                return JsonResponse({'error': 'La cantidad debe ser mayor a 0'}, status=400)
            
            if tipo_movimiento not in ['entrada', 'salida']:
                return JsonResponse({'error': 'Tipo de movimiento inválido'}, status=400)
            
            # Obtener objetos
            try:
                producto = Producto.objects.get(id=producto_id, activo=True)
                bodega = Bodega.objects.get(id=bodega_id, activa=True)
            except (Producto.DoesNotExist, Bodega.DoesNotExist):
                return JsonResponse({'error': 'Producto o bodega no encontrados'}, status=404)
            
            # Realizar ajuste en transacción
            with transaction.atomic():
                # Obtener o crear stock
                stock, created = Stock.objects.get_or_create(
                    producto=producto,
                    bodega=bodega,
                    defaults={'cantidad': 0}
                )
                
                stock_anterior = stock.cantidad
                
                # Aplicar ajuste según el tipo
                if tipo_movimiento == 'entrada':
                    stock.cantidad += cantidad
                elif tipo_movimiento == 'salida':
                    if stock.cantidad < cantidad:
                        return JsonResponse({
                            'error': f'Stock insuficiente. Disponible: {stock.cantidad}, solicitado: {cantidad}'
                        }, status=400)
                    stock.cantidad -= cantidad
                
                stock.save()
                
                # Crear movimiento de inventario
                movimiento = MovimientoInventario.objects.create(
                    producto=producto,
                    bodega=bodega,
                    tipo_movimiento='ajuste',  # Usar 'ajuste' para todos los ajustes de inventario
                    cantidad=cantidad if tipo_movimiento == 'entrada' else -cantidad,  # Cantidad negativa para salidas
                    costo_unitario=producto.costo_promedio or Decimal('0.00'),
                    motivo=motivo,
                    observaciones=f'{motivo_form or f"Ajuste de {tipo_movimiento}"}. Stock anterior: {stock_anterior}, Stock actual: {stock.cantidad}. {observaciones}',
                    usuario=request.user
                )
                
                return JsonResponse({
                    'success': True,
                    'message': f'Ajuste de {tipo_movimiento} realizado exitosamente',
                    'movimiento_id': str(movimiento.id),
                    'stock_anterior': stock_anterior,
                    'stock_actual': stock.cantidad,
                    'cantidad_ajustada': cantidad,
                    'pdf_url': f'/inventario/ajustes/{movimiento.id}/pdf/'
                })
                
        except ValueError as e:
            return JsonResponse({'error': f'Error en los datos: {str(e)}'}, status=400)
        except Exception as e:
            return JsonResponse({'error': f'Error interno: {str(e)}'}, status=500)
    
    return JsonResponse({'error': 'Método no permitido'}, status=405)


def generar_pdf_ajuste(request, movimiento_id):
    """Vista para generar PDF del documento de ajuste de inventario"""
    movimiento = get_object_or_404(MovimientoInventario, id=movimiento_id, tipo_movimiento='ajuste')
    
    # Crear respuesta PDF
    response = HttpResponse(content_type='application/pdf')
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    tipo = 'entrada' if movimiento.cantidad > 0 else 'salida'
    filename = f'ajuste_{tipo}_{movimiento.producto.codigo}_{timestamp}.pdf'
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    # Crear PDF con reportlab
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    
    # Configurar documento
    doc = SimpleDocTemplate(response, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
    styles = getSampleStyleSheet()
    story = []
    
    # Estilo personalizado para título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1,  # Center
        textColor=colors.navy
    )
    
    # Estilo para subtítulos
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Heading2'],
        fontSize=14,
        spaceAfter=12,
        textColor=colors.darkblue
    )
    
    # Determinar tipo de ajuste
    tipo_ajuste = "ENTRADA" if movimiento.cantidad > 0 else "SALIDA"
    color_tipo = colors.green if movimiento.cantidad > 0 else colors.red
    
    # Título del documento
    story.append(Paragraph(f"DOCUMENTO DE AJUSTE DE INVENTARIO - {tipo_ajuste}", title_style))
    story.append(Paragraph("DistribucioneShaddai", styles['Normal']))
    story.append(Spacer(1, 20))
    
    # Información general del ajuste
    story.append(Paragraph("INFORMACIÓN GENERAL", subtitle_style))
    
    # Extraer stock de las observaciones si está disponible
    observaciones = movimiento.observaciones or ''
    stock_anterior = 'N/A'
    stock_actual = 'N/A'
    
    if 'Stock anterior:' in observaciones and 'Stock actual:' in observaciones:
        try:
            import re
            anterior_match = re.search(r'Stock anterior: (\d+)', observaciones)
            actual_match = re.search(r'Stock actual: (\d+)', observaciones)
            if anterior_match:
                stock_anterior = anterior_match.group(1)
            if actual_match:
                stock_actual = actual_match.group(1)
        except:
            pass
    
    ajuste_data = [
        ['Fecha del Ajuste:', movimiento.fecha_movimiento.strftime('%d/%m/%Y %H:%M')],
        ['N° de Documento:', str(movimiento.id)[:8].upper()],
        ['Tipo de Ajuste:', tipo_ajuste],
        ['Usuario Responsable:', movimiento.usuario.get_full_name() or movimiento.usuario.username],
        ['Motivo:', movimiento.get_motivo_display()],
        ['', ''],  # Línea vacía
    ]
    
    ajuste_table = Table(ajuste_data, colWidths=[2*inch, 3*inch])
    ajuste_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -2), 1, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    
    story.append(ajuste_table)
    story.append(Spacer(1, 20))
    
    # Información de la bodega
    story.append(Paragraph("INFORMACIÓN DE BODEGA", subtitle_style))
    
    bodega_data = [
        ['Bodega:', movimiento.bodega.nombre],
        ['Dirección:', movimiento.bodega.direccion or 'Sin dirección especificada'],
        ['Teléfono:', movimiento.bodega.telefono or 'No especificado'],
        ['Es Principal:', 'Sí' if movimiento.bodega.es_principal else 'No'],
    ]
    
    bodega_table = Table(bodega_data, colWidths=[1.5*inch, 3.5*inch])
    bodega_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    
    story.append(bodega_table)
    story.append(Spacer(1, 20))
    
    # Detalle del producto y movimiento
    story.append(Paragraph("DETALLE DEL MOVIMIENTO", subtitle_style))
    
    producto_data = [
        ['INFORMACIÓN DEL PRODUCTO', '', ''],
        ['Código:', movimiento.producto.codigo, ''],
        ['Nombre:', movimiento.producto.nombre, ''],
        ['Categoría:', movimiento.producto.categoria.nombre if movimiento.producto.categoria else 'Sin categoría', ''],
        ['', '', ''],
        ['MOVIMIENTO DE STOCK', '', ''],
        ['Stock Anterior:', stock_anterior, 'unidades'],
        ['Cantidad Ajustada:', f'{abs(movimiento.cantidad)}', f'unidades ({tipo_ajuste.lower()})'],
        ['Stock Resultante:', stock_actual, 'unidades'],
        ['Costo Unitario:', f'${movimiento.costo_unitario:,.2f}', 'COP'],
        ['Valor Total:', f'${abs(movimiento.cantidad) * movimiento.costo_unitario:,.2f}', 'COP'],
    ]
    
    producto_table = Table(producto_data, colWidths=[2*inch, 2*inch, 1*inch])
    producto_table.setStyle(TableStyle([
        # Encabezados
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
        ('FONTNAME', (0, 5), (-1, 5), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 5), (-1, 5), 11),
        ('BACKGROUND', (0, 5), (-1, 5), colors.lightgreen if movimiento.cantidad > 0 else colors.lightcoral),
        # Contenido
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),  # Primera columna en negrita
        # Bordes
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        # Resaltar cantidad ajustada
        ('BACKGROUND', (0, 7), (-1, 7), colors.lightyellow),
    ]))
    
    story.append(producto_table)
    story.append(Spacer(1, 20))
    
    # Observaciones
    if observaciones:
        story.append(Paragraph("OBSERVACIONES", subtitle_style))
        # Limpiar las observaciones para mostrar solo lo relevante
        obs_clean = observaciones.replace(f'Stock anterior: {stock_anterior}, Stock actual: {stock_actual}.', '').strip()
        if obs_clean:
            obs_paragraph = Paragraph(obs_clean, styles['Normal'])
            story.append(obs_paragraph)
            story.append(Spacer(1, 20))
    
    # Firmas
    story.append(Spacer(1, 30))
    story.append(Paragraph("FIRMAS Y AUTORIZACIONES", subtitle_style))
    
    firma_data = [
        ['', '', ''],
        ['_________________________', '_________________________', '_________________________'],
        ['Responsable del Ajuste', 'Supervisor de Inventario', 'Jefe de Bodega'],
        [f'{movimiento.usuario.get_full_name() or movimiento.usuario.username}', '', ''],
        ['', '', ''],
        ['Fecha: _______________', 'Fecha: _______________', 'Fecha: _______________'],
    ]
    
    firma_table = Table(firma_data, colWidths=[1.8*inch, 1.8*inch, 1.8*inch])
    firma_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('FONTNAME', (0, 2), (-1, 2), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    story.append(firma_table)
    
    # Pie de página
    story.append(Spacer(1, 20))
    footer_text = f"Documento generado automáticamente - {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    story.append(Paragraph(footer_text, styles['Normal']))
    
    # Generar PDF
    doc.build(story)
    return response


# Vistas de Recomendaciones Inteligentes
class RecomendacionesListView(AdminInventarioMixin, ListView):
    """Vista de lista de recomendaciones de reposición"""
    model = RecomendacionReposicion
    template_name = 'inventario/recomendaciones_list.html'
    context_object_name = 'recomendaciones'
    paginate_by = 25
    
    def get_queryset(self):
        queryset = RecomendacionReposicion.objects.filter(
            activa=True
        ).select_related(
            'producto', 
            'proveedor_sugerido'
        ).order_by(
            'prioridad', 
            '-fecha_generacion'
        )
        
        # Filtros
        prioridad = self.request.GET.get('prioridad')
        if prioridad:
            queryset = queryset.filter(prioridad=prioridad)
        
        estado = self.request.GET.get('estado')  
        if estado:
            queryset = queryset.filter(estado=estado)
        
        categoria = self.request.GET.get('categoria')
        if categoria:
            queryset = queryset.filter(producto__categoria_id=categoria)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Estadísticas de recomendaciones
        total_recomendaciones = self.get_queryset().count()
        
        context['stats'] = {
            'total': total_recomendaciones,
            'criticas': self.get_queryset().filter(prioridad='critica').count(),
            'altas': self.get_queryset().filter(prioridad='alta').count(), 
            'medias': self.get_queryset().filter(prioridad='media').count(),
            'bajas': self.get_queryset().filter(prioridad='baja').count(),
        }
        
        # Inversión total sugerida
        context['inversion_total'] = self.get_queryset().filter(
            valor_total_sugerido__isnull=False
        ).aggregate(
            total=Sum('valor_total_sugerido')
        )['total'] or 0
        
        # Opciones para filtros
        context['prioridades'] = RecomendacionReposicion.PRIORIDAD_CHOICES
        context['estados'] = RecomendacionReposicion.ESTADO_CHOICES
        context['categorias'] = Categoria.objects.filter(activa=True)
        
        # Filtros actuales
        context['filtro_prioridad'] = self.request.GET.get('prioridad', '')
        context['filtro_estado'] = self.request.GET.get('estado', '')
        context['filtro_categoria'] = self.request.GET.get('categoria', '')
        
        return context


class RecomendacionDetailView(AdminInventarioMixin, DetailView):
    """Vista de detalle de una recomendación"""
    model = RecomendacionReposicion
    template_name = 'inventario/recomendacion_detail.html'
    context_object_name = 'recomendacion'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Análisis adicional del producto
        producto = self.object.producto
        context['patron_ventas'] = producto.analizar_patron_ventas(30)
        
        # Proveedores alternativos
        context['proveedores_alternativos'] = ProductoProveedor.objects.filter(
            producto=producto,
            activo=True,
            disponible=True,
            proveedor__activo=True
        ).exclude(
            proveedor=self.object.proveedor_sugerido
        ).select_related('proveedor').order_by('precio_compra')
        
        return context


@csrf_exempt
def procesar_recomendacion(request, pk):
    """Vista para procesar una recomendación (aprobar, rechazar, modificar)"""
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'No autenticado'}, status=401)
    
    if not (request.user.is_superuser or 
            request.user.role in ['superadmin', 'administrador']):
        return JsonResponse({'error': 'Sin permisos'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    recomendacion = get_object_or_404(RecomendacionReposicion, pk=pk, activa=True)
    accion = request.POST.get('accion')
    notas = request.POST.get('notas', '')
    
    try:
        if accion == 'aprobar':
            # Marcar como procesando
            recomendacion.marcar_procesada(request.user, notas)
            
            mensaje = f"Recomendación aprobada y marcada para procesamiento"
            
            return JsonResponse({
                'success': True,
                'message': mensaje,
                'nuevo_estado': recomendacion.get_estado_display()
            })
            
        elif accion == 'rechazar':
            razon = request.POST.get('razon', 'Sin especificar')
            recomendacion.marcar_rechazada(request.user, razon)
            
            return JsonResponse({
                'success': True,
                'message': 'Recomendación rechazada',
                'nuevo_estado': recomendacion.get_estado_display()
            })
            
        elif accion == 'completar':
            recomendacion.marcar_completada(request.user, notas)
            
            return JsonResponse({
                'success': True,
                'message': 'Recomendación marcada como completada',
                'nuevo_estado': recomendacion.get_estado_display()
            })
            
        elif accion == 'modificar':
            nueva_cantidad = request.POST.get('cantidad')
            if nueva_cantidad:
                try:
                    nueva_cantidad = int(nueva_cantidad)
                    if nueva_cantidad > 0:
                        recomendacion.cantidad_sugerida = nueva_cantidad
                        
                        # Recalcular valor total si hay precio
                        if recomendacion.precio_sugerido:
                            recomendacion.valor_total_sugerido = (
                                recomendacion.precio_sugerido * nueva_cantidad
                            )
                        
                        recomendacion.notas_procesamiento = (
                            f"Cantidad modificada a {nueva_cantidad} unidades. {notas}"
                        )
                        recomendacion.save()
                        
                        return JsonResponse({
                            'success': True,
                            'message': f'Cantidad modificada a {nueva_cantidad} unidades',
                            'nueva_cantidad': nueva_cantidad,
                            'nuevo_valor': float(recomendacion.valor_total_sugerido) if recomendacion.valor_total_sugerido else None
                        })
                    else:
                        return JsonResponse({'error': 'La cantidad debe ser mayor a 0'}, status=400)
                except ValueError:
                    return JsonResponse({'error': 'Cantidad inválida'}, status=400)
            else:
                return JsonResponse({'error': 'Cantidad requerida'}, status=400)
        else:
            return JsonResponse({'error': 'Acción no válida'}, status=400)
            
    except Exception as e:
        return JsonResponse({'error': f'Error al procesar: {str(e)}'}, status=500)


@csrf_exempt 
def generar_recomendaciones_ajax(request):
    """Vista AJAX para generar recomendaciones en tiempo real"""
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'No autenticado'}, status=401)
    
    if not (request.user.is_superuser or 
            request.user.role in ['superadmin', 'administrador']):
        return JsonResponse({'error': 'Sin permisos'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    try:
        modo = request.POST.get('modo', 'todos')  # 'todos', 'criticos', 'especifico'
        producto_id = request.POST.get('producto_id')
        forzar = request.POST.get('forzar') == 'true'
        
        recomendaciones_generadas = 0
        productos_analizados = 0
        
        if modo == 'especifico' and producto_id:
            # Analizar producto específico
            try:
                producto = Producto.objects.get(id=producto_id, activo=True)
                productos_analizados = 1
                
                recomendacion = producto.generar_recomendacion_inteligente(forzar=forzar)
                if recomendacion:
                    recomendaciones_generadas = 1
                    
            except Producto.DoesNotExist:
                return JsonResponse({'error': 'Producto no encontrado'}, status=404)
                
        elif modo == 'criticos':
            # Analizar solo productos con alerta de stock
            productos = Producto.productos_con_alerta_stock()[:50]  # Limitar a 50
            
            for producto in productos:
                productos_analizados += 1
                recomendacion = producto.generar_recomendacion_inteligente(forzar=forzar)
                if recomendacion:
                    recomendaciones_generadas += 1
                    
        else:
            # Analizar todos los productos (limitado para evitar timeout)
            productos = Producto.objects.filter(activo=True)[:100]  # Limitar a 100
            
            for producto in productos:
                productos_analizados += 1
                recomendacion = producto.generar_recomendacion_inteligente(forzar=forzar)
                if recomendacion:
                    recomendaciones_generadas += 1
        
        return JsonResponse({
            'success': True,
            'message': f'Análisis completado: {recomendaciones_generadas} nuevas recomendaciones',
            'productos_analizados': productos_analizados,
            'recomendaciones_generadas': recomendaciones_generadas
        })
        
    except Exception as e:
        return JsonResponse({'error': f'Error en el análisis: {str(e)}'}, status=500)


class DashboardRecomendacionesView(AdminInventarioMixin, TemplateView):
    """Dashboard de recomendaciones con estadísticas y gráficos"""
    template_name = 'inventario/dashboard_recomendaciones.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Recomendaciones activas
        recomendaciones_activas = RecomendacionReposicion.objects.filter(activa=True)
        
        # Estadísticas generales
        context['estadisticas'] = {
            'total_activas': recomendaciones_activas.count(),
            'criticas': recomendaciones_activas.filter(prioridad='critica').count(),
            'urgentes': recomendaciones_activas.filter(prioridad__in=['critica', 'alta']).count(),
            'procesando': recomendaciones_activas.filter(estado='procesando').count(),
            'pendientes': recomendaciones_activas.filter(estado='pendiente').count(),
        }
        
        # Top 10 productos más urgentes
        context['productos_urgentes'] = recomendaciones_activas.filter(
            prioridad__in=['critica', 'alta']
        ).select_related('producto').order_by(
            'prioridad', 'dias_cobertura_actual'
        )[:10]
        
        # Inversión requerida por prioridad
        context['inversion_por_prioridad'] = {}
        for prioridad, _ in RecomendacionReposicion.PRIORIDAD_CHOICES:
            total = recomendaciones_activas.filter(
                prioridad=prioridad,
                valor_total_sugerido__isnull=False
            ).aggregate(total=Sum('valor_total_sugerido'))['total'] or 0
            
            context['inversion_por_prioridad'][prioridad] = {
                'total': total,
                'count': recomendaciones_activas.filter(prioridad=prioridad).count()
            }
        
        # Productos sin stock (críticos)
        context['productos_sin_stock'] = recomendaciones_activas.filter(
            stock_actual=0
        ).select_related('producto').count()
        
        # Tendencias (productos con crecimiento fuerte)
        context['productos_crecimiento'] = recomendaciones_activas.filter(
            tendencia_porcentaje__gte=20
        ).count()
        
        return context

# API Views para AJAX
from django.http import JsonResponse

def obtener_presentaciones_proveedor(request, proveedor_id, producto_id):
    """Vista API para obtener presentaciones disponibles de un proveedor para un producto específico"""
    try:
        # Verificar que el usuario tenga permisos
        if not request.user.is_authenticated:
            return JsonResponse({'error': 'No autorizado'}, status=401)
        
        # Obtener la relación producto-proveedor
        try:
            producto_proveedor = ProductoProveedor.objects.get(
                id=proveedor_id,
                producto_id=producto_id
            )
        except ProductoProveedor.DoesNotExist:
            return JsonResponse({'error': 'Relación producto-proveedor no encontrada'}, status=404)
        
        # Obtener presentaciones disponibles para este proveedor
        presentaciones = producto_proveedor.presentaciones_disponibles.filter(
            disponible=True
        ).select_related('presentacion_base').order_by('es_presentacion_preferida', 'presentacion_base__orden')
        
        presentaciones_data = []
        for pres_prov in presentaciones:
            presentaciones_data.append({
                'id': pres_prov.id,
                'nombre': pres_prov.presentacion_base.nombre,
                'unidades_por_presentacion': pres_prov.presentacion_base.unidades_por_presentacion,
                'precio_compra_presentacion': float(pres_prov.precio_compra_presentacion),
                'es_presentacion_preferida': pres_prov.es_presentacion_preferida,
                'codigo_proveedor': pres_prov.codigo_proveedor,
                'tiempo_entrega_dias': pres_prov.tiempo_entrega_dias,
                'descuento_volumen': float(pres_prov.descuento_volumen) if pres_prov.descuento_volumen else 0,
                'cantidad_descuento': pres_prov.cantidad_descuento
            })
        
        return JsonResponse({
            'presentaciones': presentaciones_data,
            'proveedor': producto_proveedor.proveedor.nombre,
            'producto': producto_proveedor.producto.codigo
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

# inventario/productos_views.py
"""
Vistas especializadas para gestión de productos
Extraído de inventario/views.py para mejorar organización
"""
from django.shortcuts import render, get_object_or_404, redirect
from django.views.generic import ListView, CreateView, DetailView, UpdateView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.urls import reverse_lazy
from django.db.models import Q, Sum, Count
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
from decimal import Decimal

from .models import Producto, Categoria, Subcategoria, Stock, Bodega
from .forms import ProductoFilterForm, ProductoForm


class InventarioViewMixin(UserPassesTestMixin):
    """Mixin para permitir ver inventario (incluye bodega)"""
    def test_func(self):
        return (self.request.user.can_adjust_inventory() or 
                self.request.user.can_view_inventory())


class AdminOnlyMixin(UserPassesTestMixin):
    """Mixin para funciones que solo puede usar el administrador"""
    def test_func(self):
        return self.request.user.is_superuser or self.request.user.groups.filter(name='admin').exists()


# ============= VISTAS DE PRODUCTOS =============

class ProductoListView(InventarioViewMixin, ListView):
    """Lista de productos con filtros avanzados"""
    model = Producto
    template_name = 'inventario/producto_list.html'
    context_object_name = 'productos'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = Producto.objects.select_related('categoria', 'subcategoria').prefetch_related('stock__bodega')
        
        # Aplicar filtros
        form = ProductoFilterForm(self.request.GET)
        if form.is_valid():
            # Filtro por texto
            search = form.cleaned_data.get('search')
            if search:
                queryset = queryset.filter(
                    Q(codigo__icontains=search) |
                    Q(nombre__icontains=search) |
                    Q(codigo_barras__icontains=search)
                )
            
            # Filtro por categoría
            categoria = form.cleaned_data.get('categoria')
            if categoria:
                queryset = queryset.filter(categoria=categoria)
            
            # Filtro por subcategoría
            subcategoria = form.cleaned_data.get('subcategoria')
            if subcategoria:
                queryset = queryset.filter(subcategoria=subcategoria)
            
            # Filtro por rango de precios
            precio_min = form.cleaned_data.get('precio_min')
            if precio_min:
                queryset = queryset.filter(precio_minorista__gte=precio_min)
            
            precio_max = form.cleaned_data.get('precio_max')
            if precio_max:
                queryset = queryset.filter(precio_minorista__lte=precio_max)
            
            # Filtro por estado activo
            activo = form.cleaned_data.get('activo')
            if activo:
                queryset = queryset.filter(activo=(activo == 'True'))
            
            # Filtro por stock
            stock_status = form.cleaned_data.get('stock_status')
            if stock_status:
                if stock_status == 'sin_stock':
                    # Productos sin stock
                    productos_sin_stock = []
                    for producto in queryset:
                        stock_total = producto.stock.aggregate(total=Sum('cantidad'))['total'] or 0
                        if stock_total == 0:
                            productos_sin_stock.append(producto.id)
                    queryset = queryset.filter(id__in=productos_sin_stock)
                
                elif stock_status == 'bajo_minimo':
                    # Productos bajo el mínimo
                    productos_bajo_minimo = []
                    for producto in queryset:
                        stock_total = producto.stock.aggregate(total=Sum('cantidad'))['total'] or 0
                        if stock_total <= producto.stock_minimo:
                            productos_bajo_minimo.append(producto.id)
                    queryset = queryset.filter(id__in=productos_bajo_minimo)
                
                elif stock_status == 'disponible':
                    # Productos con stock disponible
                    productos_con_stock = []
                    for producto in queryset:
                        stock_total = producto.stock.aggregate(total=Sum('cantidad'))['total'] or 0
                        if stock_total > 0:
                            productos_con_stock.append(producto.id)
                    queryset = queryset.filter(id__in=productos_con_stock)
        
        return queryset.order_by('codigo')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter_form'] = ProductoFilterForm(self.request.GET)
        
        # Estadísticas adicionales
        context['total_productos'] = Producto.objects.count()
        context['productos_activos'] = Producto.objects.filter(activo=True).count()
        context['productos_bajo_minimo'] = self._get_productos_bajo_minimo_count()
        
        return context
    
    def _get_productos_bajo_minimo_count(self):
        """Contar productos bajo stock mínimo"""
        count = 0
        productos = Producto.objects.filter(activo=True)
        for producto in productos:
            stock_total = producto.stock.aggregate(total=Sum('cantidad'))['total'] or 0
            if stock_total <= producto.stock_minimo:
                count += 1
        return count


class ProductoCreateView(AdminOnlyMixin, CreateView):
    """Crear nuevo producto"""
    model = Producto
    form_class = ProductoForm
    template_name = 'inventario/producto_form.html'
    success_url = reverse_lazy('inventario:producto_list')
    
    def form_valid(self, form):
        form.instance.usuario_creacion = self.request.user
        
        # Calcular precios basados en porcentajes si se proporcionaron
        costo = form.cleaned_data.get('costo_promedio', Decimal('0'))
        margen_min = form.cleaned_data.get('margen_minorista')
        margen_may = form.cleaned_data.get('margen_mayorista')
        
        if costo and margen_min is not None:
            precio_minorista = costo * (1 + (margen_min / Decimal('100')))
            form.instance.precio_minorista = precio_minorista.quantize(Decimal('0.01'))
        
        if costo and margen_may is not None:
            precio_mayorista = costo * (1 + (margen_may / Decimal('100')))
            form.instance.precio_mayorista = precio_mayorista.quantize(Decimal('0.01'))
        
        messages.success(self.request, f'Producto "{form.instance.nombre}" creado exitosamente.')
        return super().form_valid(form)


class ProductoDetailView(InventarioViewMixin, DetailView):
    """Detalle de producto con información de stock"""
    model = Producto
    template_name = 'inventario/producto_detail.html'
    context_object_name = 'producto'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Obtener stock por bodegas
        context['stocks'] = Stock.objects.filter(
            producto=self.object
        ).select_related('bodega', 'variante')
        
        # Calcular stock total
        context['stock_total'] = context['stocks'].aggregate(
            total=Sum('cantidad')
        )['total'] or 0
        
        # Histórico de movimientos recientes
        from .models import MovimientoInventario
        context['movimientos_recientes'] = MovimientoInventario.objects.filter(
            producto=self.object
        ).order_by('-fecha_movimiento')[:10]
        
        # Información de proveedores
        from .models import ProductoProveedor
        context['proveedores'] = ProductoProveedor.objects.filter(
            producto=self.object
        ).select_related('proveedor')
        
        # Estadísticas del producto
        context['necesita_reposicion'] = context['stock_total'] <= self.object.stock_minimo
        context['valor_inventario'] = context['stock_total'] * self.object.costo_promedio
        
        return context


class ProductoUpdateView(AdminOnlyMixin, UpdateView):
    """Editar producto"""
    model = Producto
    form_class = ProductoForm
    template_name = 'inventario/producto_form.html'
    success_url = reverse_lazy('inventario:producto_list')
    
    def form_valid(self, form):
        # Calcular precios basados en porcentajes si se proporcionaron
        costo = form.cleaned_data.get('costo_promedio', Decimal('0'))
        margen_min = form.cleaned_data.get('margen_minorista')
        margen_may = form.cleaned_data.get('margen_mayorista')
        
        if costo and margen_min is not None:
            precio_minorista = costo * (1 + (margen_min / Decimal('100')))
            form.instance.precio_minorista = precio_minorista.quantize(Decimal('0.01'))
        
        if costo and margen_may is not None:
            precio_mayorista = costo * (1 + (margen_may / Decimal('100')))
            form.instance.precio_mayorista = precio_mayorista.quantize(Decimal('0.01'))
        
        messages.success(self.request, f'Producto "{form.instance.nombre}" actualizado exitosamente.')
        return super().form_valid(form)


class ProductoDeleteView(AdminOnlyMixin, DeleteView):
    """Eliminar producto"""
    model = Producto
    template_name = 'inventario/producto_confirm_delete.html'
    success_url = reverse_lazy('inventario:producto_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Eliminar Producto'
        return context

    def delete(self, request, *args, **kwargs):
        try:
            producto = self.get_object()
            messages.success(request, f'Producto "{producto.nombre}" eliminado exitosamente.')
            return super().delete(request, *args, **kwargs)
        except Exception as e:
            messages.error(
                request, 
                f'No se puede eliminar el producto: {str(e)}'
            )
            return redirect('inventario:producto_editar', pk=self.get_object().id)


# ============= FUNCIONES DE PRODUCTOS =============

def exportar_productos_excel(request):
    """Vista para exportar productos a Excel"""
    if not request.user.is_authenticated:
        return HttpResponse("No autorizado", status=401)
    
    if not (request.user.can_adjust_inventory() or request.user.can_view_inventory()):
        return HttpResponse("Sin permisos", status=403)
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"
    
    # Configurar estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Definir headers basándose en permisos del usuario
    if request.user.can_see_costs():
        headers = [
            'Código', 'Nombre', 'Categoría', 'Subcategoría', 
            'Costo Promedio', 'Precio Minorista', 'Precio Mayorista',
            'Stock Total', 'Stock Mínimo', 'Estado', 'Fecha Creación'
        ]
    else:
        headers = [
            'Código', 'Nombre', 'Categoría', 'Subcategoría', 
            'Precio Minorista', 'Precio Mayorista',
            'Stock Total', 'Stock Mínimo', 'Estado', 'Fecha Creación'
        ]
    
    # Escribir headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Obtener productos con los mismos filtros que la vista de lista
    productos_view = ProductoListView()
    productos_view.request = request
    productos = productos_view.get_queryset()
    
    # Escribir datos
    row_num = 2
    for producto in productos:
        # Calcular stock total
        stock_total = producto.stock.aggregate(total=Sum('cantidad'))['total'] or 0
        
        if request.user.can_see_costs():
            row_data = [
                producto.codigo,
                producto.nombre,
                producto.categoria.nombre if producto.categoria else '',
                producto.subcategoria.nombre if producto.subcategoria else '',
                float(producto.costo_promedio),
                float(producto.precio_minorista),
                float(producto.precio_mayorista),
                stock_total,
                producto.stock_minimo,
                'Activo' if producto.activo else 'Inactivo',
                producto.fecha_creacion.strftime('%d/%m/%Y %H:%M')
            ]
        else:
            row_data = [
                producto.codigo,
                producto.nombre,
                producto.categoria.nombre if producto.categoria else '',
                producto.subcategoria.nombre if producto.subcategoria else '',
                float(producto.precio_minorista),
                float(producto.precio_mayorista),
                stock_total,
                producto.stock_minimo,
                'Activo' if producto.activo else 'Inactivo',
                producto.fecha_creacion.strftime('%d/%m/%Y %H:%M')
            ]
        
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)
        
        row_num += 1
    
    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'productos_reyes_{timestamp}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    # Guardar workbook en la respuesta
    wb.save(response)
    
    return response


def duplicar_producto(request, pk):
    """Duplicar un producto existente"""
    if not request.user.is_superuser:
        messages.error(request, 'No tienes permisos para duplicar productos.')
        return redirect('inventario:producto_list')
    
    producto_original = get_object_or_404(Producto, pk=pk)
    
    # Crear copia del producto
    producto_nuevo = Producto.objects.create(
        codigo=f"{producto_original.codigo}_COPY",
        nombre=f"{producto_original.nombre} (Copia)",
        descripcion=producto_original.descripcion,
        categoria=producto_original.categoria,
        subcategoria=producto_original.subcategoria,
        costo_promedio=producto_original.costo_promedio,
        precio_minorista=producto_original.precio_minorista,
        precio_mayorista=producto_original.precio_mayorista,
        stock_minimo=producto_original.stock_minimo,
        unidad_medida=producto_original.unidad_medida,
        usuario_creacion=request.user,
        activo=False  # Crear inactivo para que sea revisado
    )
    
    messages.success(request, f'Producto duplicado como "{producto_nuevo.nombre}". Revisa y activa cuando esté listo.')
    return redirect('inventario:producto_detail', pk=producto_nuevo.pk)


# ============= APIs DE PRODUCTOS =============

def buscar_productos_api(request):
    """API para buscar productos (autocompletado)"""
    if not request.user.can_view_inventory():
        return JsonResponse({'error': 'Sin permisos'}, status=403)
    
    term = request.GET.get('q', '') or request.GET.get('term', '')
    if len(term) < 2:
        return JsonResponse([], safe=False)
    
    productos = Producto.objects.filter(
        Q(nombre__icontains=term) |
        Q(codigo__icontains=term),
        activo=True
    )[:10]
    
    results = []
    for producto in productos:
        stock_total = producto.stock.aggregate(total=Sum('cantidad'))['total'] or 0
        
        results.append({
            'id': producto.id,
            'label': f"{producto.nombre} - {producto.codigo}",
            'value': producto.nombre,
            'codigo': producto.codigo,
            'precio_minorista': float(producto.precio_minorista),
            'precio_mayorista': float(producto.precio_mayorista),
            'stock_total': stock_total,
            'categoria': producto.categoria.nombre if producto.categoria else ''
        })
    
    return JsonResponse(results, safe=False)


def verificar_codigo_producto(request):
    """API para verificar si un código de producto ya existe"""
    if not request.user.can_adjust_inventory():
        return JsonResponse({'error': 'Sin permisos'}, status=403)
    
    codigo = request.GET.get('codigo', '')
    producto_id = request.GET.get('producto_id')  # Para edición
    
    if not codigo:
        return JsonResponse({'disponible': True})
    
    # Verificar existencia
    query = Producto.objects.filter(codigo=codigo)
    if producto_id:
        query = query.exclude(id=producto_id)
    
    existe = query.exists()
    
    return JsonResponse({
        'disponible': not existe,
        'mensaje': 'Código no disponible' if existe else 'Código disponible'
    })


def obtener_precios_producto(request, pk):
    """API para obtener precios de un producto"""
    if not request.user.can_view_inventory():
        return JsonResponse({'error': 'Sin permisos'}, status=403)
    
    try:
        producto = Producto.objects.get(pk=pk, activo=True)
        stock_total = producto.stock.aggregate(total=Sum('cantidad'))['total'] or 0
        
        data = {
            'precio_minorista': float(producto.precio_minorista),
            'precio_mayorista': float(producto.precio_mayorista),
            'stock_total': stock_total,
            'disponible': stock_total > 0
        }
        
        if request.user.can_see_costs():
            data['costo_promedio'] = float(producto.costo_promedio)
        
        return JsonResponse(data)
    except Producto.DoesNotExist:
        return JsonResponse({'error': 'Producto no encontrado'}, status=404)
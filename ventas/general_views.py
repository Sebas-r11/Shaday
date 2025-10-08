# ventas/general_views.py
"""
Vistas generales del módulo de ventas
Dashboard, búsquedas generales y funcionalidades compartidas
"""
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.http import JsonResponse
from django.db.models import Q, Sum, Count
from django.utils import timezone
from datetime import datetime, timedelta

from .models import Cliente, Cotizacion, Pedido, Factura, Entrega
from inventario.models import Producto


class VentasRequiredMixin(UserPassesTestMixin):
    """Mixin para verificar permisos de ventas"""
    def test_func(self):
        return self.request.user.can_create_sales()


# ============= DASHBOARD =============

@login_required
def dashboard_view(request):
    """Dashboard principal de ventas"""
    if not request.user.can_create_sales():
        return redirect('dashboard')
    
    # Estadísticas del mes actual
    hoy = timezone.now().date()
    primer_dia_mes = hoy.replace(day=1)
    
    # Ventas del mes
    ventas_mes = Factura.objects.filter(
        fecha_creacion__date__gte=primer_dia_mes,
        estado='pagada'
    ).aggregate(Sum('total'))['total__sum'] or 0
    
    # Pedidos pendientes
    pedidos_pendientes = Pedido.objects.filter(estado='pendiente').count()
    pedidos_alistamiento = Pedido.objects.filter(estado='proceso').count()
    
    # Facturas por cobrar
    facturas_pendientes = Factura.objects.filter(
        estado='pendiente'
    ).aggregate(Sum('total'))['total__sum'] or 0
    
    # Entregas del día
    entregas_hoy = Entrega.objects.filter(fecha_programada__date=hoy).count()
    entregas_pendientes = Entrega.objects.filter(
        estado__in=['pendiente', 'asignada', 'en_ruta']
    ).count()
    
    # Cotizaciones pendientes
    cotizaciones_pendientes = Cotizacion.objects.filter(estado='pendiente').count()
    
    # Ventas últimos 7 días
    hace_7_dias = hoy - timedelta(days=7)
    ventas_semana = []
    
    for i in range(7):
        fecha = hace_7_dias + timedelta(days=i)
        ventas_dia = Factura.objects.filter(
            fecha_creacion__date=fecha,
            estado='pagada'
        ).aggregate(Sum('total'))['total__sum'] or 0
        
        ventas_semana.append({
            'fecha': fecha.strftime('%d/%m'),
            'ventas': float(ventas_dia)
        })
    
    # Top productos vendidos del mes
    from inventario.models import MovimientoInventario
    top_productos = MovimientoInventario.objects.filter(
        fecha_movimiento__date__gte=primer_dia_mes,
        tipo_movimiento='salida',
        motivo='venta'
    ).values(
        'producto__nombre'
    ).annotate(
        cantidad_vendida=Sum('cantidad')
    ).order_by('-cantidad_vendida')[:5]
    
    # Clientes con más compras del mes
    top_clientes = Factura.objects.filter(
        fecha_creacion__date__gte=primer_dia_mes,
        estado='pagada'
    ).values(
        'cliente__nombre_completo'
    ).annotate(
        total_compras=Sum('total'),
        num_facturas=Count('id')
    ).order_by('-total_compras')[:5]
    
    context = {
        'ventas_mes': ventas_mes,
        'pedidos_pendientes': pedidos_pendientes,
        'pedidos_alistamiento': pedidos_alistamiento,
        'facturas_pendientes': facturas_pendientes,
        'entregas_hoy': entregas_hoy,
        'entregas_pendientes': entregas_pendientes,
        'cotizaciones_pendientes': cotizaciones_pendientes,
        'ventas_semana': ventas_semana,
        'top_productos': top_productos,
        'top_clientes': top_clientes,
        'title': 'Dashboard de Ventas'
    }
    
    return render(request, 'ventas/dashboard.html', context)

@login_required
def dashboard_charts_view(request):
    """Dashboard avanzado con gráficos interactivos"""
    if not request.user.can_create_sales():
        return redirect('dashboard')
    context = {
        'ventas_mes': ventas_mes,
        'pedidos_pendientes': pedidos_pendientes,
        'pedidos_alistamiento': pedidos_alistamiento,
        'facturas_pendientes': facturas_pendientes,
        'entregas_hoy': entregas_hoy,
        'entregas_pendientes': entregas_pendientes,
        'cotizaciones_pendientes': cotizaciones_pendientes,
        'ventas_semana': ventas_semana,
        'top_productos': top_productos,
        'top_clientes': top_clientes,
        'title': 'Dashboard de Ventas'
    }
    
    return render(request, 'ventas/dashboard.html', context)

@login_required
def dashboard_charts_view(request):
    """Dashboard avanzado con gráficos interactivos"""
    if not request.user.can_create_sales():
        return redirect('dashboard')
    
    context = {
        'title': 'Dashboard Avanzado'
    }
    
    return render(request, 'ventas/dashboard_charts.html', context)


# ============= APIS GENERALES =============

@login_required
def buscar_productos_api(request):
    """API para buscar productos (usado en autocompletado)"""
    if not request.user.can_create_sales():
        return JsonResponse({'error': 'Sin permisos'}, status=403)
    
    # Aceptar tanto 'q' como 'term' para compatibilidad
    term = request.GET.get('q', '') or request.GET.get('term', '')
    if len(term) < 2:
        return JsonResponse({'productos': []})
    
    # Verificar si Producto está importado
    try:
        from inventario.models import Producto
    except ImportError:
        return JsonResponse({'error': 'Modelo Producto no disponible'}, status=500)
    
    productos = Producto.objects.filter(
        Q(nombre__icontains=term) |
        Q(codigo__icontains=term),
        activo=True
    )[:10]
    
    results = []
    for producto in productos:
        # Obtener stock total de todas las bodegas
        stock_total = sum(stock.cantidad for stock in producto.stock.all()) if hasattr(producto, 'stock') else 0
        
        results.append({
            'id': producto.id,
            'nombre': producto.nombre,
            'codigo': producto.codigo,
            'precio': float(producto.precio_minorista or 0),
            'precio_minorista': float(producto.precio_minorista or 0),
            'precio_mayorista': float(producto.precio_mayorista or 0),
            'stock': stock_total,
            'categoria': producto.categoria.nombre if producto.categoria else ''
        })
    
    return JsonResponse({'productos': results})


@login_required
def obtener_precio_producto(request, producto_id):
    """API para obtener precio actual de un producto"""
    if not request.user.can_create_sales():
        return JsonResponse({'error': 'Sin permisos'}, status=403)
    
    try:
        producto = Producto.objects.get(id=producto_id, activo=True)
        return JsonResponse({
            'precio': float(producto.precio),
            'stock': producto.stock,
            'disponible': producto.stock > 0
        })
    except Producto.DoesNotExist:
        return JsonResponse({'error': 'Producto no encontrado'}, status=404)


@login_required
def verificar_stock_disponible(request):
    """API para verificar stock de múltiples productos"""
    if not request.user.can_create_sales():
        return JsonResponse({'error': 'Sin permisos'}, status=403)
    
    import json
    
    if request.method == 'POST':
        data = json.loads(request.body)
        productos_verificar = data.get('productos', [])
        
        results = []
        for item in productos_verificar:
            producto_id = item.get('producto_id')
            cantidad_solicitada = item.get('cantidad', 0)
            
            try:
                producto = Producto.objects.get(id=producto_id, activo=True)
                disponible = producto.stock >= cantidad_solicitada
                
                results.append({
                    'producto_id': producto_id,
                    'nombre': producto.nombre,
                    'stock_actual': producto.stock,
                    'cantidad_solicitada': cantidad_solicitada,
                    'disponible': disponible,
                    'faltante': max(0, cantidad_solicitada - producto.stock)
                })
            except Producto.DoesNotExist:
                results.append({
                    'producto_id': producto_id,
                    'error': 'Producto no encontrado'
                })
        
        return JsonResponse({
            'productos': results,
            'todos_disponibles': all(r.get('disponible', False) for r in results if 'error' not in r)
        })
    
    return JsonResponse({'error': 'Método no permitido'}, status=405)


# ============= FUNCIONES DE UTILIDAD =============

def calcular_estadisticas_periodo(fecha_inicio, fecha_fin):
    """Función para calcular estadísticas de un período específico"""
    stats = {}
    
    # Ventas del período
    facturas_periodo = Factura.objects.filter(
        fecha_creacion__date__gte=fecha_inicio,
        fecha_creacion__date__lte=fecha_fin,
        estado='pagada'
    )
    
    stats['total_ventas'] = facturas_periodo.aggregate(Sum('total'))['total__sum'] or 0
    stats['numero_facturas'] = facturas_periodo.count()
    stats['promedio_venta'] = stats['total_ventas'] / stats['numero_facturas'] if stats['numero_facturas'] > 0 else 0
    
    # Pedidos del período
    pedidos_periodo = Pedido.objects.filter(
        fecha_creacion__date__gte=fecha_inicio,
        fecha_creacion__date__lte=fecha_fin
    )
    
    stats['total_pedidos'] = pedidos_periodo.count()
    stats['pedidos_completados'] = pedidos_periodo.filter(estado='completado').count()
    stats['tasa_conversion'] = (stats['pedidos_completados'] / stats['total_pedidos'] * 100) if stats['total_pedidos'] > 0 else 0
    
    # Entregas del período
    entregas_periodo = Entrega.objects.filter(
        fecha_creacion__date__gte=fecha_inicio,
        fecha_creacion__date__lte=fecha_fin
    )
    
    stats['total_entregas'] = entregas_periodo.count()
    stats['entregas_exitosas'] = entregas_periodo.filter(estado='entregada').count()
    stats['tasa_entrega'] = (stats['entregas_exitosas'] / stats['total_entregas'] * 100) if stats['total_entregas'] > 0 else 0
    
    return stats


# ============= APIS PARA GRÁFICOS DEL DASHBOARD =============

@login_required
def api_ventas_por_mes(request):
    """API para gráfico de ventas por mes de los últimos 12 meses"""
    from django.db.models import Sum
    from datetime import datetime, timedelta
    import calendar
    
    hoy = timezone.now().date()
    hace_12_meses = hoy - timedelta(days=365)
    
    # Obtener ventas por mes
    ventas_por_mes = []
    for i in range(12):
        # Calcular el mes
        mes_actual = hoy.replace(day=1) - timedelta(days=30*i)
        mes_siguiente = mes_actual.replace(day=28) + timedelta(days=4)
        mes_siguiente = mes_siguiente.replace(day=1)
        
        ventas_mes = Factura.objects.filter(
            fecha_creacion__date__gte=mes_actual,
            fecha_creacion__date__lt=mes_siguiente,
            estado='pagada'
        ).aggregate(Sum('total'))['total__sum'] or 0
        
        ventas_por_mes.append({
            'mes': calendar.month_name[mes_actual.month][:3] + f' {mes_actual.year}',
            'ventas': float(ventas_mes),
            'fecha': mes_actual.isoformat()
        })
    
    ventas_por_mes.reverse()  # Ordenar cronológicamente
    
    return JsonResponse({
        'success': True,
        'data': ventas_por_mes
    })

@login_required  
def api_productos_mas_vendidos(request):
    """API para gráfico de productos más vendidos"""
    from inventario.models import MovimientoInventario
    from django.db.models import Sum
    
    hoy = timezone.now().date()
    hace_30_dias = hoy - timedelta(days=30)
    
    productos_vendidos = MovimientoInventario.objects.filter(
        fecha_movimiento__date__gte=hace_30_dias,
        tipo_movimiento='salida',
        motivo='venta'
    ).values(
        'producto__nombre'
    ).annotate(
        cantidad_vendida=Sum('cantidad')
    ).order_by('-cantidad_vendida')[:10]
    
    data = []
    for item in productos_vendidos:
        data.append({
            'producto': item['producto__nombre'],
            'cantidad': float(item['cantidad_vendida'])
        })
    
    return JsonResponse({
        'success': True,
        'data': data
    })

@login_required
def api_estados_pedidos(request):
    """API para gráfico de distribución de estados de pedidos"""
    estados_count = {}
    
    for estado_code, estado_name in Pedido.ESTADO_CHOICES:
        count = Pedido.objects.filter(estado=estado_code).count()
        if count > 0:
            estados_count[estado_name] = count
    
    data = []
    for estado, count in estados_count.items():
        data.append({
            'estado': estado,
            'cantidad': count
        })
    
    return JsonResponse({
        'success': True,
        'data': data
    })

@login_required
def api_ventas_por_vendedor(request):
    """API para gráfico de ventas por vendedor del mes"""
    from django.contrib.auth import get_user_model
    
    User = get_user_model()
    hoy = timezone.now().date()
    primer_dia_mes = hoy.replace(day=1)
    
    # Obtener vendedores con ventas
    vendedores_ventas = Factura.objects.filter(
        fecha_creacion__date__gte=primer_dia_mes,
        estado='pagada',
        cliente__vendedor_asignado__isnull=False
    ).values(
        'cliente__vendedor_asignado__first_name',
        'cliente__vendedor_asignado__last_name'
    ).annotate(
        total_ventas=Sum('total'),
        num_ventas=Count('id')
    ).order_by('-total_ventas')[:8]
    
    data = []
    for item in vendedores_ventas:
        nombre = f"{item['cliente__vendedor_asignado__first_name']} {item['cliente__vendedor_asignado__last_name']}"
        if not nombre.strip():
            nombre = "Vendedor Sin Nombre"
            
        data.append({
            'vendedor': nombre,
            'ventas': float(item['total_ventas']),
            'num_ventas': item['num_ventas']
        })
    
    return JsonResponse({
        'success': True,
        'data': data
    })

@login_required
def api_estadisticas_dashboard(request):
    """API para estadísticas generales del dashboard"""
    hoy = timezone.now().date()
    primer_dia_mes = hoy.replace(day=1)
    mes_anterior = (primer_dia_mes - timedelta(days=1)).replace(day=1)
    
    # Estadísticas del mes actual
    pedidos_mes = Pedido.objects.filter(fecha_creacion__date__gte=primer_dia_mes).count()
    ventas_mes = Factura.objects.filter(
        fecha_creacion__date__gte=primer_dia_mes,
        estado='pagada'
    ).aggregate(Sum('total'))['total__sum'] or 0
    
    # Estadísticas del mes anterior para comparación
    pedidos_mes_anterior = Pedido.objects.filter(
        fecha_creacion__date__gte=mes_anterior,
        fecha_creacion__date__lt=primer_dia_mes
    ).count()
    
    ventas_mes_anterior = Factura.objects.filter(
        fecha_creacion__date__gte=mes_anterior,
        fecha_creacion__date__lt=primer_dia_mes,
        estado='pagada'
    ).aggregate(Sum('total'))['total__sum'] or 0
    
    # Calcular crecimientos
    crecimiento_pedidos = 0
    if pedidos_mes_anterior > 0:
        crecimiento_pedidos = ((pedidos_mes - pedidos_mes_anterior) / pedidos_mes_anterior) * 100
    
    crecimiento_ventas = 0
    if ventas_mes_anterior > 0:
        crecimiento_ventas = ((ventas_mes - ventas_mes_anterior) / ventas_mes_anterior) * 100
    
    # Pedidos pendientes
    pedidos_pendientes = Pedido.objects.filter(estado='pendiente').count()
    pedidos_proceso = Pedido.objects.filter(estado='proceso').count()
    
    # Alertas de stock
    try:
        from inventario.models import AlertaStock
        alertas_stock = AlertaStock.objects.filter(vista=False).count()
    except:
        alertas_stock = 0
    
    return JsonResponse({
        'success': True,
        'data': {
            'pedidos_mes': pedidos_mes,
            'ventas_mes': float(ventas_mes),
            'crecimiento_pedidos': round(crecimiento_pedidos, 1),
            'crecimiento_ventas': round(crecimiento_ventas, 1),
            'pedidos_pendientes': pedidos_pendientes,
            'pedidos_proceso': pedidos_proceso,
            'alertas_stock': alertas_stock,
            'mes_actual': primer_dia_mes.strftime('%B %Y')
        }
    })

# ============= VISTAS DE PRUEBA/DESARROLLO =============

@login_required
def test_autocompletado(request):
    """Vista de prueba para autocompletado"""
    return render(request, 'ventas/test_autocompletado.html')


@login_required
def test_ajax_simple(request):
    """Vista de prueba para AJAX"""
    if request.method == 'POST':
        return JsonResponse({'success': True, 'message': 'AJAX funcionando correctamente'})
    return render(request, 'ventas/test_ajax.html')


@login_required
def test_syntax(request):
    """Vista de prueba para verificar sintaxis"""
    return JsonResponse({'status': 'ok', 'message': 'Sintaxis correcta'})


# ============= SISTEMA DE REPORTES DETALLADOS =============

@login_required
def reportes_view(request):
    """Vista principal del sistema de reportes"""
    if not request.user.can_create_sales():
        return redirect('dashboard')
    
    context = {
        'titulo': 'Sistema de Reportes',
        'puede_ventas': request.user.can_create_sales(),
        'puede_inventario': request.user.can_adjust_inventory(),
        'puede_compras': request.user.can_create_sales(),  # Usar can_create_sales para compras
    }
    return render(request, 'ventas/reportes.html', context)


@login_required
def reporte_ventas(request):
    """Generar reporte detallado de ventas con filtros"""
    if not request.user.can_create_sales():
        return JsonResponse({'error': 'Sin permisos'}, status=403)
    
    # Obtener parámetros de filtro
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    cliente_id = request.GET.get('cliente')
    vendedor_id = request.GET.get('vendedor')
    estado = request.GET.get('estado')
    formato = request.GET.get('formato', 'html')
    
    # Construir query base
    queryset = Factura.objects.all()
    
    # Aplicar filtros
    if fecha_inicio:
        try:
            fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
            queryset = queryset.filter(fecha_creacion__date__gte=fecha_inicio)
        except ValueError:
            pass
    
    if fecha_fin:
        try:
            fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
            queryset = queryset.filter(fecha_creacion__date__lte=fecha_fin)
        except ValueError:
            pass
    
    if cliente_id:
        queryset = queryset.filter(cliente_id=cliente_id)
    
    if vendedor_id:
        queryset = queryset.filter(vendedor_id=vendedor_id)
    
    if estado:
        queryset = queryset.filter(estado=estado)
    
    # Ordenar por fecha más reciente
    facturas = queryset.order_by('-fecha_creacion')
    
    # Calcular totales
    total_ventas = facturas.aggregate(Sum('total'))['total__sum'] or 0
    total_facturas = facturas.count()
    promedio_factura = total_ventas / total_facturas if total_facturas > 0 else 0
    
    # Estadísticas por estado
    stats_estado = facturas.values('estado').annotate(
        cantidad=Count('id'),
        total=Sum('total')
    ).order_by('estado')
    
    # Si es export, generar archivo
    if formato in ['excel', 'csv', 'pdf']:
        return export_reporte_ventas(facturas, formato, {
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin,
            'total_ventas': total_ventas,
            'total_facturas': total_facturas
        })
    
    # Preparar datos para template
    context = {
        'titulo': 'Reporte de Ventas',
        'facturas': facturas[:100],  # Limitar a 100 para performance
        'total_ventas': total_ventas,
        'total_facturas': total_facturas,
        'promedio_factura': promedio_factura,
        'stats_estado': stats_estado,
        'filtros': {
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin,
            'cliente_id': cliente_id,
            'vendedor_id': vendedor_id,
            'estado': estado,
        },
        'clientes': Cliente.objects.filter(activo=True)[:50],
        'vendedores': request.user.__class__.objects.filter(
            groups__name='Vendedores'
        )[:20] if hasattr(request.user.__class__, 'objects') else [],
    }
    
    return render(request, 'ventas/reporte_ventas.html', context)


@login_required  
def reporte_inventario(request):
    """Generar reporte detallado de inventario"""
    if not request.user.can_adjust_inventory():
        return JsonResponse({'error': 'Sin permisos'}, status=403)
    
    from inventario.models import Producto, MovimientoInventario, AlertaStock
    
    # Parámetros de filtro
    categoria_id = request.GET.get('categoria')
    stock_bajo = request.GET.get('stock_bajo')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    formato = request.GET.get('formato', 'html')
    
    # Query base productos
    productos = Producto.objects.filter(activo=True)
    
    if categoria_id:
        productos = productos.filter(categoria_id=categoria_id)
    
    if stock_bajo == 'true':
        productos = productos.filter(stock_minimo__gt=0)  # Placeholder hasta encontrar el campo correcto
    
    # Movimientos de inventario
    movimientos = MovimientoInventario.objects.all()
    
    if fecha_inicio:
        try:
            fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
            movimientos = movimientos.filter(fecha_movimiento__date__gte=fecha_inicio)
        except ValueError:
            pass
    
    if fecha_fin:
        try:
            fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
            movimientos = movimientos.filter(fecha_movimiento__date__lte=fecha_fin)
        except ValueError:
            pass
    
    # Estadísticas generales  
    total_productos = productos.count()
    productos_stock_bajo = 0  # Placeholder - calcularemos manualmente
    valor_inventario = productos.aggregate(
        total=Sum('precio_minorista')
    )['total'] or 0
    
    # Alertas activas
    alertas_activas = AlertaStock.objects.filter(
        activa=True,
        producto__in=productos
    ).count()
    
    # Top productos por movimiento
    top_movimientos = movimientos.values(
        'producto__nombre'
    ).annotate(
        total_movimientos=Count('id'),
        cantidad_total=Sum('cantidad')
    ).order_by('-total_movimientos')[:10]
    
    if formato in ['excel', 'csv', 'pdf']:
        return export_reporte_inventario(productos, formato, {
            'total_productos': total_productos,
            'productos_stock_bajo': productos_stock_bajo,
            'valor_inventario': valor_inventario,
            'alertas_activas': alertas_activas
        })
    
    # Calcular valores totales para cada producto
    productos_con_valores = []
    for producto in productos[:100]:
        stock_actual = getattr(producto, 'stock_total', 0)
        valor_total = float(stock_actual * producto.precio_minorista)
        productos_con_valores.append({
            'producto': producto,
            'stock_actual': stock_actual,
            'valor_total': valor_total
        })

    context = {
        'titulo': 'Reporte de Inventario',
        'productos_con_valores': productos_con_valores,
        'total_productos': total_productos,
        'productos_stock_bajo': productos_stock_bajo,
        'valor_inventario': valor_inventario,
        'alertas_activas': alertas_activas,
        'top_movimientos': top_movimientos,
        'filtros': {
            'categoria_id': categoria_id,
            'stock_bajo': stock_bajo,
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin,
        }
    }
    
    return render(request, 'ventas/reporte_inventario.html', context)


@login_required
def reporte_compras(request):
    """Generar reporte detallado de compras"""
    if not request.user.can_create_sales():
        return JsonResponse({'error': 'Sin permisos'}, status=403)
    
    # Este es un placeholder - asumo que hay un módulo de compras
    # Por ahora mostraremos los pedidos como "compras internas"
    
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    estado = request.GET.get('estado')
    formato = request.GET.get('formato', 'html')
    
    queryset = Pedido.objects.all()
    
    if fecha_inicio:
        try:
            fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
            queryset = queryset.filter(fecha_creacion__date__gte=fecha_inicio)
        except ValueError:
            pass
    
    if fecha_fin:
        try:
            fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
            queryset = queryset.filter(fecha_creacion__date__lte=fecha_fin)
        except ValueError:
            pass
    
    if estado:
        queryset = queryset.filter(estado=estado)
    
    pedidos = queryset.order_by('-fecha_creacion')
    
    # Estadísticas
    total_pedidos = pedidos.count()
    total_valor = pedidos.aggregate(Sum('total'))['total__sum'] or 0
    promedio_pedido = total_valor / total_pedidos if total_pedidos > 0 else 0
    
    stats_estado = pedidos.values('estado').annotate(
        cantidad=Count('id'),
        total=Sum('total')
    ).order_by('estado')
    
    if formato in ['excel', 'csv', 'pdf']:
        return export_reporte_compras(pedidos, formato, {
            'total_pedidos': total_pedidos,
            'total_valor': total_valor,
            'promedio_pedido': promedio_pedido
        })
    
    context = {
        'titulo': 'Reporte de Compras/Pedidos',
        'pedidos': pedidos[:100],
        'total_pedidos': total_pedidos,
        'total_valor': total_valor,
        'promedio_pedido': promedio_pedido,
        'stats_estado': stats_estado,
        'filtros': {
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin,
            'estado': estado,
        }
    }
    
    return render(request, 'ventas/reporte_compras.html', context)


# ============= FUNCIONES DE EXPORT =============

def export_reporte_ventas(facturas, formato, stats):
    """Exportar reporte de ventas en diferentes formatos"""
    from django.http import HttpResponse
    import io
    
    if formato == 'excel':
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Ventas"
        
        # Headers
        headers = ['Fecha', 'Número', 'Cliente', 'Vendedor', 'Estado', 'Total']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        for row, factura in enumerate(facturas, 2):
            ws.cell(row=row, column=1, value=factura.fecha_creacion.strftime('%Y-%m-%d'))
            ws.cell(row=row, column=2, value=factura.numero_factura)
            ws.cell(row=row, column=3, value=str(factura.cliente))
            ws.cell(row=row, column=4, value=str(factura.vendedor) if factura.vendedor else 'N/A')
            ws.cell(row=row, column=5, value=factura.get_estado_display())
            ws.cell(row=row, column=6, value=float(factura.total))
        
        # Totales al final
        total_row = len(facturas) + 3
        ws.cell(row=total_row, column=5, value="TOTAL:").font = Font(bold=True)
        ws.cell(row=total_row, column=6, value=float(stats['total_ventas'])).font = Font(bold=True)
        
        # Preparar respuesta
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="reporte_ventas.xlsx"'
        return response
    
    elif formato == 'csv':
        import csv
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="reporte_ventas.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Fecha', 'Número', 'Cliente', 'Vendedor', 'Estado', 'Total'])
        
        for factura in facturas:
            writer.writerow([
                factura.fecha_creacion.strftime('%Y-%m-%d'),
                factura.numero_factura,
                str(factura.cliente),
                str(factura.vendedor) if factura.vendedor else 'N/A',
                factura.get_estado_display(),
                float(factura.total)
            ])
        
        return response
    
    elif formato == 'pdf':
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="reporte_ventas.pdf"'
        
        doc = SimpleDocTemplate(response, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Título
        title = Paragraph("Reporte de Ventas", styles['Title'])
        story.append(title)
        story.append(Spacer(1, 20))
        
        # Resumen
        summary = Paragraph(f"Total de Facturas: {stats['total_facturas']}<br/>Valor Total: ${stats['total_ventas']:,.2f}", styles['Normal'])
        story.append(summary)
        story.append(Spacer(1, 20))
        
        # Tabla de datos
        data = [['Fecha', 'Número', 'Cliente', 'Estado', 'Total']]
        for factura in facturas[:50]:  # Limitar para PDF
            data.append([
                factura.fecha_creacion.strftime('%Y-%m-%d'),
                factura.numero_factura,
                str(factura.cliente)[:30],  # Truncar nombres largos
                factura.get_estado_display(),
                f"${factura.total:,.2f}"
            ])
        
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(table)
        doc.build(story)
        return response


def export_reporte_inventario(productos, formato, stats):
    """Exportar reporte de inventario"""
    from django.http import HttpResponse
    import io
    
    if formato == 'excel':
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Inventario"
        
        headers = ['Código', 'Nombre', 'Categoría', 'Stock', 'Precio', 'Valor Total']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        for row, producto in enumerate(productos, 2):
            ws.cell(row=row, column=1, value=producto.codigo)
            ws.cell(row=row, column=2, value=producto.nombre)
            ws.cell(row=row, column=3, value=str(producto.categoria) if producto.categoria else 'N/A')
            ws.cell(row=row, column=4, value=0)  # Stock placeholder
            ws.cell(row=row, column=5, value=float(producto.precio_minorista))
            ws.cell(row=row, column=6, value=float(0 * producto.precio_minorista))
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="reporte_inventario.xlsx"'
        return response
    
    elif formato == 'csv':
        import csv
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="reporte_inventario.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Código', 'Nombre', 'Categoría', 'Stock', 'Precio', 'Valor Total'])
        
        for producto in productos:
            writer.writerow([
                producto.codigo,
                producto.nombre,
                str(producto.categoria) if producto.categoria else 'N/A',
                0,  # Stock placeholder
                float(producto.precio_minorista),
                float(0 * producto.precio_minorista)
            ])
        
        return response
    
    # Si no es excel o csv, retornar respuesta por defecto
    return HttpResponse("Formato no soportado", status=400)


def export_reporte_compras(pedidos, formato, stats):
    """Exportar reporte de compras/pedidos"""
    from django.http import HttpResponse
    import io
    
    if formato == 'excel':
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Compras"
        
        headers = ['Fecha', 'Número', 'Cliente', 'Estado', 'Total']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        for row, pedido in enumerate(pedidos, 2):
            ws.cell(row=row, column=1, value=pedido.fecha_creacion.strftime('%Y-%m-%d'))
            ws.cell(row=row, column=2, value=pedido.numero)
            ws.cell(row=row, column=3, value=str(pedido.cliente))
            ws.cell(row=row, column=4, value=pedido.get_estado_display())
            ws.cell(row=row, column=5, value=float(pedido.total))
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="reporte_compras.xlsx"'
        return response
    
    elif formato == 'csv':
        import csv
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="reporte_compras.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Fecha', 'Número', 'Cliente', 'Estado', 'Total'])
        
        for pedido in pedidos:
            writer.writerow([
                pedido.fecha_creacion.strftime('%Y-%m-%d'),
                pedido.numero,
                str(pedido.cliente),
                pedido.get_estado_display(),
                float(pedido.total)
            ])
        
        return response
    
    # Si no es excel o csv, retornar respuesta por defecto
    return HttpResponse("Formato no soportado", status=400)